# -*- coding: utf-8 -*-
"""
Importa colaboradores e férias reais da planilha Teknisa Retail
"""
import openpyxl
from datetime import datetime, timedelta
from excel_handler import ExcelHandler
from models import Colaborador, Ferias

CAMINHO_FONTE = r'C:\Users\Teknisa\Downloads\TEKNISA - CONTROLE DE FÉRIAS  RETAIL.xlsx'

def parse_data(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d")
    except:
        return None

def importar():
    print("Lendo planilha fonte...")
    wb = openpyxl.load_workbook(CAMINHO_FONTE, data_only=True)

    # --- Coletar colaboradores únicos das abas ---
    colaboradores_raw = {}  # nome -> {admis, tipo, uf}

    for aba in ['FÉRIAS', 'FÉRIAS 2025', 'FÉRIAS 2026']:
        ws = wb[aba]
        header_row = None
        for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if r[0] == 'NOME':
                break
            header_row = None
        # Dados começam na linha 4 (FÉRIAS) ou 4 (2025/2026)
        start = 4 if aba == 'FÉRIAS' else 4
        for row in ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True):
            nome = str(row[0]).strip() if row[0] else None
            if not nome or nome in ('None', 'NOME'):
                continue
            admis = parse_data(row[3])
            tipo = str(row[1]).strip() if row[1] else 'Consultor'
            uf = str(row[2]).strip() if row[2] else ''
            if nome not in colaboradores_raw:
                colaboradores_raw[nome] = {'admis': admis, 'tipo': tipo, 'uf': uf}
            elif admis and not colaboradores_raw[nome]['admis']:
                colaboradores_raw[nome]['admis'] = admis

    print(f"  Colaboradores encontrados: {len(colaboradores_raw)}")

    # --- Montar lista de Colaborador ---
    handler = ExcelHandler()
    colaboradores = []
    for i, (nome, info) in enumerate(sorted(colaboradores_raw.items()), start=1):
        admis = info['admis'] or datetime(2020, 1, 1)
        time_str = f"{info['tipo']} - {info['uf']}" if info['uf'] else info['tipo']
        c = Colaborador(i, nome, admis, time_str, ativo=True)
        colaboradores.append(c)
        print(f"  [{i}] {nome} | admissão: {admis.strftime('%d/%m/%Y')} | {time_str}")

    handler.salvar_colaboradores(colaboradores)

    # --- Coletar férias planejadas (2026) e realizadas (2025) ---
    colab_por_nome = {c.nome.strip(): c for c in colaboradores}
    ferias_planejadas = []
    ferias_realizadas_list = []
    fid = 1

    mapa_abas = {
        'FÉRIAS 2025': 'Realizado',
        'FÉRIAS 2026': 'Confirmado',
        'FÉRIAS':      'Confirmado',
    }

    for aba, status_base in mapa_abas.items():
        ws = wb[aba]
        start = 4 if aba == 'FÉRIAS' else 4
        nome_atual = None
        for row in ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True):
            cell_nome = str(row[0]).strip() if row[0] else None
            if cell_nome and cell_nome not in ('None',):
                nome_atual = cell_nome

            if not nome_atual:
                continue

            inicio = parse_data(row[8])
            dias_val = row[9]
            status_cell = str(row[11]).strip() if row[11] else ''

            if not inicio or not dias_val:
                continue

            dias = int(float(dias_val))
            fim = inicio + timedelta(days=dias - 1)
            status = 'Realizado' if (status_base == 'Realizado' and status_cell in ('CONFIRMADA', 'CONCLUIDA', 'CONCLUÍDA')) else \
                     'Confirmado' if status_cell == 'CONFIRMADA' else 'Planejado'

            colab = colab_por_nome.get(nome_atual)
            if not colab:
                continue

            f = Ferias(fid, colab.id, inicio, fim, status)
            fid += 1

            if status == 'Realizado':
                ferias_realizadas_list.append((f, colab.nome))
            else:
                ferias_planejadas.append(f)

    # Remover duplicatas (mesma pessoa, mesmas datas)
    vistos = set()
    ferias_unicas = []
    for f in ferias_planejadas:
        chave = (f.colaborador_id, f.data_inicio.date(), f.data_fim.date())
        if chave not in vistos:
            vistos.add(chave)
            ferias_unicas.append(f)

    handler.salvar_ferias_planejadas(ferias_unicas)

    # Salvar realizadas
    wb_dest = openpyxl.load_workbook(handler.caminho)
    ws_real = wb_dest['Férias Realizadas']
    ws_real.delete_rows(2, ws_real.max_row)
    for idx, (f, nome) in enumerate(ferias_realizadas_list, start=1):
        ws_real.append([idx, f.colaborador_id, nome,
                        f.data_inicio.strftime('%Y-%m-%d'),
                        f.data_fim.strftime('%Y-%m-%d'),
                        f.dias, f.data_inicio.year])
    wb_dest.save(handler.caminho)
    wb_dest.close()

    print(f"\n  Férias planejadas importadas: {len(ferias_unicas)}")
    print(f"  Férias realizadas importadas: {len(ferias_realizadas_list)}")
    print("\nImportação concluída!")

if __name__ == '__main__':
    importar()
