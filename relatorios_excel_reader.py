import pandas as pd
import openpyxl
import os
import re

REQUIRED_SHEETS = ['Tradicionais', 'Rollout', 'Treinamento']


def column_letter_to_index(col_letter):
    """Converte letra de coluna Excel (ex: 'KZ') para índice numérico (0-based)"""
    col_letter = col_letter.upper()
    index = 0
    for i, char in enumerate(reversed(col_letter)):
        index += (ord(char) - ord('A') + 1) * (26 ** i)
    return index - 1  # Converte para 0-based index


def get_period_from_sheet(filepath):
    wb = openpyxl.load_workbook(filepath)
    first_sheet = wb.sheetnames[0]
    return first_sheet


def _extract_period(filepath, sheet_name):
    """Tenta extrair o período (YYYY-MM) do nome do arquivo; usa nome da aba como fallback."""
    filename = os.path.basename(filepath)
    # Procura padrão YYYY-MM ou YYYY_MM no nome do arquivo
    m = re.search(r'(\d{4})[-_](\d{2})', filename)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # Fallback: nome da primeira aba
    return sheet_name


def load_workbook(filepath):
    try:
        xls = pd.ExcelFile(filepath)
        sheets = xls.sheet_names

        first_sheet = sheets[0]
        period = _extract_period(filepath, first_sheet)

        dataframes = {}
        dataframes[period] = pd.read_excel(filepath, sheet_name=first_sheet)

        for sheet in REQUIRED_SHEETS:
            if sheet in sheets:
                df = pd.read_excel(filepath, sheet_name=sheet)
                dataframes[sheet] = df
            else:
                print(f"Aviso: Aba '{sheet}' nao encontrada na planilha")

        return dataframes, period
    except Exception as e:
        print(f"Erro ao carregar arquivo: {e}")
        raise


def extract_coordinators(dataframes):
    coordinators = set()

    if 'Base projetos' in dataframes or list(dataframes.keys())[0] in dataframes:
        period_key = list(dataframes.keys())[0]
        df_base = dataframes[period_key]

        if 'Gerente do Projeto' in df_base.columns:
            values = df_base['Gerente do Projeto'].dropna().unique()
            coordinators.update(values)

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in dataframes:
            continue

        df = dataframes[sheet_name]
        first_col = df.columns[0]

        if df.shape[0] > 2:
            values = df.iloc[2:, 0].dropna().unique()
            coordinators.update(values)

    return sorted(list(coordinators))


def suggest_commission_column(df, sheet_name):
    """Analisa o dataframe e sugere qual coluna provavelmente contém comissão do mês

    Retorna: dict com informações da coluna sugerida
    """

    if df.empty or df.shape[0] < 1:
        return None

    # Procura por "COMISSÃO DO MÊS" nas colunas
    for col_idx, col_header in enumerate(df.columns):
        if col_header and 'COMISSÃO' in str(col_header).upper() and 'MÊS' in str(col_header).upper():
            # Encontrou! Agora procura por "Total" próximo
            # Procura na primeira linha de dados
            if df.shape[0] > 0:
                for search_idx in range(col_idx, min(col_idx + 10, df.shape[1])):
                    search_val = str(df.iloc[0, search_idx]).upper() if pd.notna(df.iloc[0, search_idx]) else ""
                    if 'TOTAL' in search_val:
                        return _build_column_info(df, search_idx, sheet_name)

    # Se não encontrou "COMISSÃO DO MÊS", procura nas primeiras linhas
    for row_idx in range(min(2, df.shape[0])):
        for col_idx in range(df.shape[1]):
            cell_val = str(df.iloc[row_idx, col_idx]).upper() if pd.notna(df.iloc[row_idx, col_idx]) else ""
            if 'COMISSÃO' in cell_val and 'MÊS' in cell_val:
                # Procura por "Total" na mesma linha
                for search_idx in range(col_idx, min(col_idx + 10, df.shape[1])):
                    search_val = str(df.iloc[row_idx, search_idx]).upper() if pd.notna(df.iloc[row_idx, search_idx]) else ""
                    if 'TOTAL' in search_val:
                        return _build_column_info(df, search_idx, sheet_name)

    return None


def _build_column_info(df, col_index, sheet_name):
    """Constrói dict com informações da coluna"""

    if col_index >= df.shape[1]:
        return None

    # Converte índice para letra
    col_letter = _index_to_column_letter(col_index)

    # Pega o header
    header = df.iloc[0, col_index] if df.shape[0] > 0 else "N/A"

    # Conta valores não-zero
    if df.shape[0] > 2:
        non_zero_count = sum(1 for val in df.iloc[2:, col_index] if pd.notna(val) and val != 0)
    else:
        non_zero_count = 0

    return {
        'sheet_name': sheet_name,
        'col_letter': col_letter,
        'col_index': col_index,
        'header': header,
        'non_zero_count': non_zero_count,
        'total_rows': max(0, df.shape[0] - 2)
    }


def _index_to_column_letter(col_index):
    """Converte índice numérico (0-based) para letra de coluna Excel"""
    col_index = col_index + 1  # Converte para 1-based
    result = ""
    while col_index > 0:
        col_index -= 1
        result = chr(65 + (col_index % 26)) + result
        col_index //= 26
    return result


def get_commission_columns_suggestions(dataframes):
    """Retorna sugestões de colunas de comissão para cada aba"""
    suggestions = {}

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in dataframes:
            continue

        df = dataframes[sheet_name]
        suggestion = suggest_commission_column(df, sheet_name)
        if suggestion:
            suggestions[sheet_name] = suggestion

    return suggestions


def filter_by_coordinator(dataframes, coordinator_name, commission_columns=None):
    """Filtra dados por coordenador

    Args:
        dataframes: Dict com dataframes de cada aba
        coordinator_name: Nome do coordenador
        commission_columns: Dict com as colunas escolhidas {sheet_name: col_index}
    """
    filtered_data = {}

    period_key = list(dataframes.keys())[0]

    df_period = dataframes[period_key]
    if 'Gerente do Projeto' in df_period.columns:
        filtered_data[period_key] = df_period[
            df_period['Gerente do Projeto'] == coordinator_name
        ].reset_index(drop=True)
    else:
        filtered_data[period_key] = df_period

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in dataframes:
            continue

        df = dataframes[sheet_name]

        if df.shape[0] > 2:
            mask = df.iloc[2:, 0] == coordinator_name
            header_rows = df.iloc[:2]
            data_rows = df.iloc[list(mask.index[mask])]
            filtered_data[sheet_name] = pd.concat([header_rows, data_rows]).reset_index(drop=True)
        else:
            filtered_data[sheet_name] = df

    return filtered_data
