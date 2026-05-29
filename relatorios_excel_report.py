from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os
from datetime import datetime


def _safe_filepath(output_path, filename):
    """Retorna caminho disponível; adiciona timestamp se o arquivo estiver aberto/bloqueado."""
    filepath = os.path.join(output_path, filename)
    if not os.path.exists(filepath):
        return filepath, filename
    try:
        with open(filepath, 'a'):
            pass
        return filepath, filename
    except PermissionError:
        base, ext = os.path.splitext(filename)
        ts = datetime.now().strftime("%H%M%S")
        new_name = f"{base}_{ts}{ext}"
        return os.path.join(output_path, new_name), new_name


def _find_doc_mes_col(df):
    """Localiza a coluna 'DOCUMENTOS DO MÊS' / 'Doc. no Mês?'."""
    for col_idx, col_name in enumerate(df.columns):
        s = str(col_name).upper().replace('\n', ' ')
        if 'DOCUMENTOS' in s:
            return col_idx
    return None


def _has_doc_mes(row_data, doc_mes_col):
    """Retorna True se a linha tem Documento do Mês preenchido."""
    if doc_mes_col is None or doc_mes_col >= len(row_data):
        return True  # Coluna não encontrada: não filtra
    val = row_data.iloc[doc_mes_col]
    return pd.notna(val) and str(val).strip() != ''


def _find_delivery_cols(df, doc_mes_col):
    """Localiza colunas PP, FAP, RE, TE logo após DOCUMENTOS DO MÊS (sub-headers da linha 0)."""
    if doc_mes_col is None or df.shape[0] < 1:
        return {}
    delivery_names = {'PP', 'FAP', 'RE', 'TE'}
    delivery_cols = {}
    for col_idx in range(doc_mes_col + 1, min(doc_mes_col + 6, df.shape[1])):
        sub_header = str(df.iloc[0, col_idx]).strip().upper()
        if sub_header in delivery_names and sub_header not in delivery_cols:
            delivery_cols[sub_header] = col_idx
    return delivery_cols


def _get_delivery_counts(row_data, delivery_cols):
    """Retorna dict {nome: quantidade} para cada tipo de entrega encontrado."""
    counts = {}
    for name in ['PP', 'FAP', 'RE', 'TE']:
        if name not in delivery_cols:
            continue
        val = row_data.iloc[delivery_cols[name]]
        if pd.notna(val) and val is not False:
            try:
                counts[name] = int(float(val))
            except (TypeError, ValueError):
                counts[name] = 0
        else:
            counts[name] = 0
    return counts


def create_final_report(filtered_data, coordinator_name, period, output_path, commission_columns=None):
    """Cria relatório com colunas de entregas (PP/FAP/RE/TE) e comissão do mês."""

    if commission_columns is None:
        commission_columns = {}

    base_filename = f"Relatorio_{coordinator_name}_{period}.xlsx"
    filepath, filename = _safe_filepath(output_path, base_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    header_fill    = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font    = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    delivery_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    delivery_font  = Font(name='Arial', size=11, bold=True, color="375623")
    section_fill   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font   = Font(name='Arial', size=10, bold=True)
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    title_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    title_font  = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    label_fill  = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    label_font  = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    value_font  = Font(name='Arial', size=11, color="1F3864")
    meta_border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'),  bottom=Side(style='medium')
    )

    row_num = 1

    # Linha 1 — título
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    c = ws.cell(row_num, 1)
    c.value     = "RELATÓRIO DE COMISSIONAMENTO"
    c.font      = title_font
    c.fill      = title_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = meta_border
    ws.row_dimensions[row_num].height = 28
    row_num += 1

    # Linha 2 — Coordenador
    ws.cell(row_num, 1).value     = "COORDENADOR"
    ws.cell(row_num, 1).font      = label_font
    ws.cell(row_num, 1).fill      = label_fill
    ws.cell(row_num, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row_num, 1).border    = meta_border
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=6)
    c3 = ws.cell(row_num, 2)
    c3.value     = coordinator_name
    c3.font      = Font(name='Arial', size=12, bold=True, color="1F3864")
    c3.fill      = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    c3.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c3.border    = meta_border
    ws.row_dimensions[row_num].height = 22
    row_num += 2

    sheet_order = [period, 'Tradicionais', 'Rollout', 'Treinamento']

    default_col_mapping = {
        'Tradicionais': 187,
        'Rollout': 47,
        'Treinamento': 75,
    }

    for sheet_name in sheet_order:
        if sheet_name not in filtered_data:
            continue

        df = filtered_data[sheet_name]
        if df.empty:
            continue

        comissao_col  = commission_columns.get(sheet_name, default_col_mapping.get(sheet_name, 187))
        doc_mes_col   = _find_doc_mes_col(df)
        delivery_cols = _find_delivery_cols(df, doc_mes_col)

        # Ordem fixa dos tipos de entrega presentes nesta aba
        delivery_names = [n for n in ['PP', 'FAP', 'RE', 'TE'] if n in delivery_cols]

        # Coleta apenas linhas com Documento do Mês preenchido (pula as 2 linhas de sub-cabeçalho)
        valid_rows = []
        for _, row_data in df.iloc[2:].iterrows():
            projeto = str(row_data.iloc[3]) if pd.notna(row_data.iloc[3]) else ""
            if not projeto:
                continue
            if not _has_doc_mes(row_data, doc_mes_col):
                continue

            valor = 0
            if len(row_data) > comissao_col:
                raw = row_data.iloc[comissao_col]
                valor = raw if pd.notna(raw) and raw is not False else 0

            counts = _get_delivery_counts(row_data, delivery_cols)
            valid_rows.append((projeto, counts, valor))

        # Omite a seção inteira se não houver linhas válidas
        if not valid_rows:
            continue

        # Número de colunas: Projeto + entregas + Comissão
        total_cols = 1 + len(delivery_names) + 1
        comissao_col_out = total_cols  # última coluna (1-based)

        # Cabeçalho da seção
        ws.cell(row_num, 1).value = sheet_name
        ws.cell(row_num, 1).font  = section_font
        ws.cell(row_num, 1).fill  = section_fill
        row_num += 1

        # Cabeçalho das colunas
        headers = ['Projeto'] + delivery_names + ['Comissao do Mes']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row_num, col_idx)
            cell.value     = header
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx == 1 or col_idx == total_cols:
                # Projeto e Comissão: azul
                cell.fill = header_fill
                cell.font = header_font
            else:
                # Colunas de entrega: verde
                cell.fill = delivery_fill
                cell.font = delivery_font
        row_num += 1

        # Linhas de dados
        for projeto, counts, valor in valid_rows:
            ws.cell(row_num, 1).value = projeto
            ws.cell(row_num, 1).border    = thin_border
            ws.cell(row_num, 1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            for i, name in enumerate(delivery_names, 2):
                cell = ws.cell(row_num, i)
                cell.value     = counts.get(name, 0)
                cell.border    = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='top')

            cell_com = ws.cell(row_num, comissao_col_out)
            cell_com.value         = valor
            cell_com.border        = thin_border
            cell_com.number_format = '#,##0.00'
            cell_com.alignment     = Alignment(horizontal='right', vertical='top')

            row_num += 1

        # Linha de total
        total_fill   = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        total_font   = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        total_border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'),  bottom=Side(style='medium')
        )
        total_valor = sum(v for _, _, v in valid_rows if isinstance(v, (int, float)))

        # Células do total (Projeto até penúltima coluna): label "TOTAL"
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=comissao_col_out - 1)
        cell_lbl = ws.cell(row_num, 1)
        cell_lbl.value     = "TOTAL"
        cell_lbl.font      = total_font
        cell_lbl.fill      = total_fill
        cell_lbl.border    = total_border
        cell_lbl.alignment = Alignment(horizontal='right', vertical='center')

        # Célula do valor total
        cell_tot = ws.cell(row_num, comissao_col_out)
        cell_tot.value         = total_valor
        cell_tot.font          = total_font
        cell_tot.fill          = total_fill
        cell_tot.border        = total_border
        cell_tot.number_format = '#,##0.00'
        cell_tot.alignment     = Alignment(horizontal='right', vertical='center')
        row_num += 1

        # Linha Comissão Gerente (60% do total)
        ger_fill   = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ger_font   = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        ger_border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='thin'),    bottom=Side(style='medium')
        )
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=comissao_col_out - 1)
        cell_gl = ws.cell(row_num, 1)
        cell_gl.value     = "Comissão Gerente (60%)"
        cell_gl.font      = ger_font
        cell_gl.fill      = ger_fill
        cell_gl.border    = ger_border
        cell_gl.alignment = Alignment(horizontal='right', vertical='center')

        cell_gv = ws.cell(row_num, comissao_col_out)
        cell_gv.value         = total_valor * 0.60
        cell_gv.font          = ger_font
        cell_gv.fill          = ger_fill
        cell_gv.border        = ger_border
        cell_gv.number_format = '#,##0.00'
        cell_gv.alignment     = Alignment(horizontal='right', vertical='center')

        row_num += 2  # linha em branco após bloco de totais

    # Larguras de coluna
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 18

    try:
        wb.save(filepath)
        return True, filename
    except Exception as e:
        return False, str(e)


# ─────────────────────────────────────────────────────────────────
# RELATÓRIO CONSOLIDADO
# ─────────────────────────────────────────────────────────────────

def create_consolidated_report(all_filtered_data, period, output_path, commission_columns=None):
    """Relatório consolidado: agrupado por coordenador, linhas = tipos de aba."""

    if commission_columns is None:
        commission_columns = {}

    base_filename = f"Relatorio_Consolidado_{period}.xlsx"
    filepath, filename = _safe_filepath(output_path, base_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    # Estilos
    title_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    title_font  = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    label_fill  = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    label_font  = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    coord_fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    coord_font  = Font(name='Arial', size=11, bold=True, color="1F3864")
    resumo_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    resumo_font = Font(name='Arial', size=10, bold=True, color="1F3864")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    deliv_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    deliv_font  = Font(name='Arial', size=10, bold=True, color="375623")
    total_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    total_font  = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    grand_fill  = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    grand_font  = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    ger_fill    = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ger_font    = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    data_font   = Font(name='Arial', size=10)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))
    med  = Border(left=Side(style='medium'), right=Side(style='medium'),
                  top=Side(style='medium'),  bottom=Side(style='medium'))
    ger_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                        top=Side(style='thin'),    bottom=Side(style='medium'))

    default_col_mapping = {'Tradicionais': 187, 'Rollout': 47, 'Treinamento': 75}
    sheet_order = ['Tradicionais', 'Rollout', 'Treinamento']

    # ── 1ª passagem: coleta dados por coordenador / aba ──────────
    # coord_sheet_data[coord][sheet] = {'counts': {...}, 'commission': float}
    # grand_summary[coord] = {'commission': float, 'PP': int, ...}
    coord_sheet_data = {}
    grand_summary    = {}
    all_del_seen     = []   # union de tipos de entrega vistos

    for coord_name, filtered_data in all_filtered_data.items():
        coord_sheet_data[coord_name] = {}
        grand_summary[coord_name] = {'commission': 0.0,
                                     'PP': 0, 'FAP': 0, 'RE': 0, 'TE': 0}

        for sheet_name in sheet_order:
            if sheet_name not in filtered_data:
                continue
            df = filtered_data[sheet_name]
            if df.empty:
                continue

            comissao_col  = commission_columns.get(sheet_name,
                            default_col_mapping.get(sheet_name, 187))
            doc_mes_col   = _find_doc_mes_col(df)
            delivery_cols = _find_delivery_cols(df, doc_mes_col)
            del_names     = [n for n in ['PP', 'FAP', 'RE', 'TE'] if n in delivery_cols]
            for n in del_names:
                if n not in all_del_seen:
                    all_del_seen.append(n)

            total_valor  = 0.0
            total_counts = {'PP': 0, 'FAP': 0, 'RE': 0, 'TE': 0}

            for _, row_data in df.iloc[2:].iterrows():
                projeto = str(row_data.iloc[3]) if pd.notna(row_data.iloc[3]) else ""
                if not projeto or not _has_doc_mes(row_data, doc_mes_col):
                    continue
                if len(row_data) > comissao_col:
                    raw = row_data.iloc[comissao_col]
                    v = raw if pd.notna(raw) and raw is not False else 0
                    if isinstance(v, (int, float)):
                        total_valor += v
                for n, cnt in _get_delivery_counts(row_data, delivery_cols).items():
                    total_counts[n] = total_counts.get(n, 0) + cnt

            coord_sheet_data[coord_name][sheet_name] = {
                'counts': total_counts,
                'commission': total_valor,
            }
            grand_summary[coord_name]['commission'] += total_valor
            for n in ['PP', 'FAP', 'RE', 'TE']:
                grand_summary[coord_name][n] += total_counts.get(n, 0)

    # Colunas de entrega disponíveis (ordem fixa)
    g_del    = [n for n in ['PP', 'FAP', 'RE', 'TE'] if n in all_del_seen]
    num_cols = 1 + len(g_del) + 1      # Aba/Coord + entregas + Comissão
    comm_col = num_cols
    NCOLS    = max(6, num_cols)

    # ── Helpers ──────────────────────────────────────────────────
    def _hdr_row(row, labels, fills, fonts):
        for ci, (lbl, f, fn) in enumerate(zip(labels, fills, fonts), 1):
            c = ws.cell(row, ci)
            c.value = lbl; c.font = fn; c.fill = f; c.border = thin
            c.alignment = Alignment(horizontal='center', vertical='center')

    def _data_row(row, values):
        for ci, v in enumerate(values, 1):
            c = ws.cell(row, ci)
            c.value = v; c.font = data_font; c.border = thin
            if ci == 1:
                c.alignment = Alignment(horizontal='left', vertical='center')
            elif ci == num_cols:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                c.alignment = Alignment(horizontal='center', vertical='center')

    def _total_row(row, label, del_totals, com_total, fill, font):
        c = ws.cell(row, 1)
        c.value = label; c.font = font; c.fill = fill
        c.border = med; c.alignment = Alignment(horizontal='right', vertical='center')
        for ci, n in enumerate(g_del, 2):
            c = ws.cell(row, ci)
            c.value = del_totals.get(n, 0)
            c.font = font; c.fill = fill; c.border = med
            c.alignment = Alignment(horizontal='center', vertical='center')
        c = ws.cell(row, num_cols)
        c.value = com_total; c.font = font; c.fill = fill
        c.border = med; c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right', vertical='center')

    def _ger_row(row, com_total):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=num_cols - 1)
        cgl = ws.cell(row, 1)
        cgl.value = "Comissão Gerente (60%)"; cgl.font = ger_font
        cgl.fill = ger_fill; cgl.border = ger_border
        cgl.alignment = Alignment(horizontal='right', vertical='center')
        cgv = ws.cell(row, num_cols)
        cgv.value = com_total * 0.60; cgv.font = ger_font
        cgv.fill = ger_fill; cgv.border = ger_border
        cgv.number_format = '#,##0.00'
        cgv.alignment = Alignment(horizontal='right', vertical='center')

    # Cabeçalhos de colunas reutilizados
    col_hdrs  = ['Aba'] + g_del + ['Comissão do Mês']
    col_fills = [header_fill] + [deliv_fill] * len(g_del) + [header_fill]
    col_fonts = [header_font] + [deliv_font] * len(g_del) + [header_font]

    row_num = 1

    # ── Cabeçalho do relatório ───────────────────────────────────
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NCOLS)
    c = ws.cell(row_num, 1)
    c.value = "RELATÓRIO CONSOLIDADO DE COMISSIONAMENTO"
    c.font = title_font; c.fill = title_fill; c.border = med
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 28
    row_num += 1

    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NCOLS)
    c = ws.cell(row_num, 1)
    c.value = f"Período: {period}   |   Coordenadores: {len(all_filtered_data)}"
    c.font = label_font; c.fill = label_fill; c.border = med
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row_num].height = 20
    row_num += 2

    # ── Seções por coordenador ───────────────────────────────────
    for coord_name in all_filtered_data:
        sheet_data = coord_sheet_data.get(coord_name, {})
        sheets_present = [s for s in sheet_order if s in sheet_data]
        if not sheets_present:
            continue

        # Cabeçalho do coordenador
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=NCOLS)
        c = ws.cell(row_num, 1)
        c.value = coord_name; c.font = coord_font; c.fill = coord_fill
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        # Cabeçalho das colunas
        _hdr_row(row_num, col_hdrs, col_fills, col_fonts)
        row_num += 1

        # Linhas por tipo de aba
        coord_com_total = 0.0
        coord_del_total = {n: 0 for n in g_del}

        for sheet_name in sheet_order:
            if sheet_name not in sheet_data:
                continue
            sd = sheet_data[sheet_name]
            values = ([sheet_name] +
                      [sd['counts'].get(n, 0) for n in g_del] +
                      [sd['commission']])
            _data_row(row_num, values)
            row_num += 1
            coord_com_total += sd['commission']
            for n in g_del:
                coord_del_total[n] += sd['counts'].get(n, 0)

        # TOTAL do coordenador
        _total_row(row_num, "TOTAL", coord_del_total, coord_com_total,
                   total_fill, total_font)
        row_num += 1

        # Comissão Gerente (60%)
        _ger_row(row_num, coord_com_total)
        row_num += 2   # linha em branco após bloco

    # ── Resumo Geral ─────────────────────────────────────────────
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NCOLS)
    c = ws.cell(row_num, 1)
    c.value = "RESUMO GERAL — TOTAL ACUMULADO POR COORDENADOR"
    c.font = resumo_font; c.fill = resumo_fill
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row_num].height = 18
    row_num += 1

    # Cabeçalho do resumo (Coordenador em vez de Aba)
    res_hdrs  = ['Coordenador'] + g_del + ['Total Comissão do Mês']
    res_fills = [header_fill] + [deliv_fill] * len(g_del) + [header_fill]
    res_fonts = [header_font] + [deliv_font] * len(g_del) + [header_font]
    _hdr_row(row_num, res_hdrs, res_fills, res_fonts)
    row_num += 1

    # Linha por coordenador no resumo
    for coord_name, data in grand_summary.items():
        values = ([coord_name] +
                  [data.get(n, 0) for n in g_del] +
                  [data['commission']])
        _data_row(row_num, values)
        row_num += 1

    # TOTAL GERAL
    gt_del = {n: sum(v.get(n, 0) for v in grand_summary.values()) for n in g_del}
    gt_com = sum(v['commission'] for v in grand_summary.values())
    _total_row(row_num, "TOTAL GERAL", gt_del, gt_com, grand_fill, grand_font)
    row_num += 1

    # Comissão Gerente (60% total geral)
    ger_font_g   = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    ger_border_g = Border(left=Side(style='medium'), right=Side(style='medium'),
                          top=Side(style='thin'),    bottom=Side(style='medium'))
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=num_cols - 1)
    cgl = ws.cell(row_num, 1)
    cgl.value = "Comissão Gerente (60%)"; cgl.font = ger_font_g
    cgl.fill = ger_fill; cgl.border = ger_border_g
    cgl.alignment = Alignment(horizontal='right', vertical='center')
    cgv = ws.cell(row_num, num_cols)
    cgv.value = gt_com * 0.60; cgv.font = ger_font_g
    cgv.fill = ger_fill; cgv.border = ger_border_g
    cgv.number_format = '#,##0.00'
    cgv.alignment = Alignment(horizontal='right', vertical='center')

    # Larguras
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 20

    try:
        wb.save(filepath)
        return True, filename
    except Exception as e:
        return False, str(e)
