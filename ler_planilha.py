# -*- coding: utf-8 -*-
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

caminho = r'C:\Users\Teknisa\Downloads\TEKNISA - CONTROLE DE FÉRIAS  RETAIL.xlsx'
wb = openpyxl.load_workbook(caminho)

print("ABAS:", wb.sheetnames)

# Aba FÉRIAS
print("\n=== ABA FÉRIAS (cols A-L, linhas 1-20) ===")
ws = wb['FÉRIAS']
for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=12, values_only=True):
    if any(v is not None for v in row):
        print(list(row))

# Aba FÉRIAS 2025
print("\n=== ABA FÉRIAS 2025 (cols A-L, linhas 1-20) ===")
ws2 = wb['FÉRIAS 2025']
for row in ws2.iter_rows(min_row=1, max_row=20, min_col=1, max_col=12, values_only=True):
    if any(v is not None for v in row):
        print(list(row))

# Aba FÉRIAS 2026
print("\n=== ABA FÉRIAS 2026 (cols A-L, linhas 1-20) ===")
ws3 = wb['FÉRIAS 2026']
for row in ws3.iter_rows(min_row=1, max_row=20, min_col=1, max_col=12, values_only=True):
    if any(v is not None for v in row):
        print(list(row))

# Aba Acompanhamento Consultores
print("\n=== ABA Acompanhamento Consultores (cols A-P, linhas 1-25) ===")
ws4 = wb['Acompanhamento Consultores']
for row in ws4.iter_rows(min_row=1, max_row=25, min_col=1, max_col=16, values_only=True):
    if any(v is not None for v in row):
        print(list(row))
