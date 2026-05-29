# -*- coding: utf-8 -*-
import os
import sys
import socket
import calendar
from datetime import datetime, timedelta, date
from functools import wraps

# ─── Converter Supabase direct URL para pooler URL (IPv4, melhor para serverless) ──
def _supabase_pooler_url(url):
    """
    Converte qualquer URL Supabase (direta ou pooler antigo) para o pooler correto.
    Pooler correto: aws-1-us-west-2 com projeto bxzwvvxhmknmgilcyqub
    """
    import re

    # Extrair senha de qualquer URL Supabase
    m = re.match(r'postgresql://[^:]+:(.+?)@', url)
    if not m:
        return url

    password = m.group(1)

    # Sempre usar o pooler correto (bxzwvvxhmknmgilcyqub / aws-1-us-west-2)
    # independente do URL original configurado
    if 'supabase.co' in url or 'supabase.com' in url:
        return (
            f'postgresql://postgres.bxzwvvxhmknmgilcyqub:{password}'
            f'@aws-1-us-west-2.pooler.supabase.com:6543/postgres?sslmode=require'
        )

    return url

from dotenv import load_dotenv
load_dotenv()

from dateutil.relativedelta import relativedelta
from flask import (Flask, render_template, request, redirect,
                   url_for, jsonify, flash, send_file)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from models import Colaborador, Ferias
from database import db, User, ColaboradorDB, FeriasDB, ERPProjetoDB, ERPModuloDB, ERPUnidadeDB, ERPAtividadeDB, ComissionamentoDB
from validators import FeriasValidator
from analytics import FeriasAnalytics
from erp_models import Projeto, Modulo, Unidade, Atividade
from erp_validators import ProjetoValidator
from erp_analytics import ProjetoAnalytics

# ─── App & config ─────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')

# Determinar qual banco usar
# Se DATABASE_URL estiver definido, usar sempre (Supabase/PostgreSQL)
# Caso contrário, usar SQLite local ou /tmp no Vercel
os.makedirs(DATA_DIR, exist_ok=True)

db_url = os.environ.get('DATABASE_URL', '')

if db_url:
    # Converter postgres:// para postgresql:// se necessário
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)
    # Adicionar sslmode=require se não tiver
    if 'postgresql://' in db_url and 'sslmode' not in db_url:
        db_url += '?sslmode=require'
    # Converter Supabase direct → pooler (resolve problema IPv6 no Vercel)
    db_url = _supabase_pooler_url(db_url)
else:
    # Usar SQLite (local ou /tmp no Vercel)
    is_vercel = os.environ.get('VERCEL') == '1'
    if is_vercel:
        db_url = 'sqlite:////tmp/ferias_data.db'
    else:
        db_url = f'sqlite:///{os.path.join(DATA_DIR, "ferias_data.db")}'

app.config['SQLALCHEMY_DATABASE_URI']        = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# ─── Flask-Login ──────────────────────────────────────────────────────────────

login_manager = LoginManager(app)
login_manager.login_view         = 'login'
login_manager.login_message      = 'Faça login para acessar o sistema.'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(uid):
    try:
        return User.query.get(int(uid))
    except:
        return None

# ─── Verificação de banco inicializado ────────────────────────────────────────

def is_db_initialized():
    """Verifica se o banco de dados foi inicializado"""
    try:
        with app.app_context():
            # Tenta contar usuários - se conseguir, o banco está pronto
            User.query.count()
            return True
    except:
        return False

@app.before_request
def check_db_init():
    """Redireciona para /init se o banco não estiver inicializado"""
    # Rotas que não precisam de banco de dados
    exempt_routes = ['init_database', 'static']

    if request.endpoint in exempt_routes:
        return

    if not is_db_initialized():
        return redirect(url_for('init_database'))

# ─── Inicialização do banco ───────────────────────────────────────────────────
# Comentado para permitir inicialização via endpoint
# with app.app_context():
#     db.create_all()
#     if User.query.count() == 0:
#         admin = User(
#             username=os.environ.get('ADMIN_USER', 'admin'),
#             nome='Gestor',
#             perfil='gestor',
#         )
#         admin.set_senha(os.environ.get('ADMIN_PASSWORD', 'admin123'))
#         db.session.add(admin)
#         db.session.commit()
#         print('[OK] Usuario gestor criado (username: admin)')

# ─── Decorators ───────────────────────────────────────────────────────────────

def gestor_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_gestor:
            flash('Acesso restrito ao gestor.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

# ─── Constantes ───────────────────────────────────────────────────────────────

CORES = [
    '#1e88e5', '#43a047', '#e53935', '#8e24aa', '#f4511e',
    '#00897b', '#3949ab', '#d81b60', '#6d4c41', '#039be5',
    '#7cb342', '#ef6c00',
]
CARGO_ORDEM = {'gerencia': 0, 'coordenador': 1, 'consultor': 2, 'estagiario': 3, 'outro': 4}
CARGO_LABEL = {
    'gerencia': 'Gerência', 'coordenador': 'Coordenadores',
    'consultor': 'Consultores', 'estagiario': 'Estagiários', 'outro': 'Outros',
}

# ─── Helpers gerais ───────────────────────────────────────────────────────────

def cor_colab(idx: int) -> str:
    return CORES[idx % len(CORES)]

def parse_cargo_uf(time_str: str):
    partes = [p.strip() for p in str(time_str or '').split('-', 1)]
    cargo_raw = partes[0].lower() if partes else ''
    uf = partes[1].upper() if len(partes) > 1 else ''
    if   'ger'    in cargo_raw: cargo = 'gerencia'
    elif 'coord'  in cargo_raw: cargo = 'coordenador'
    elif 'consul' in cargo_raw: cargo = 'consultor'
    elif 'estag'  in cargo_raw: cargo = 'estagiario'
    else:                       cargo = 'outro'
    return cargo, uf

def sort_key(colab):
    cargo, uf = parse_cargo_uf(colab.time)
    return (CARGO_ORDEM.get(cargo, 9), uf, colab.nome.lower())

def sort_colaboradores(colabs):
    return sorted(colabs, key=sort_key)

# ─── Converters DB → domínio ──────────────────────────────────────────────────

def db_to_colab(c: ColaboradorDB) -> Colaborador:
    return Colaborador(
        id=c.id,
        nome=c.nome,
        data_admissao=datetime.combine(c.data_admissao, datetime.min.time()),
        time=c.time or '',
        ativo=c.ativo,
        cidade=c.cidade or '',
    )

def db_to_ferias(f: FeriasDB) -> Ferias:
    obj = Ferias(
        id=f.id,
        colaborador_id=f.colaborador_id,
        data_inicio=datetime.combine(f.data_inicio, datetime.min.time()),
        data_fim=datetime.combine(f.data_fim, datetime.min.time()),
        status=f.status,
    )
    obj.conflito_detectado = f.conflito_detectado or False
    obj.conflito_aprovado  = f.conflito_aprovado  or False
    return obj

def db_to_projeto(p: ERPProjetoDB) -> Projeto:
    responsavel = None
    if p.responsavel_id:
        resp_colab = ColaboradorDB.query.get(p.responsavel_id)
        responsavel = resp_colab.nome if resp_colab else None

    proj = Projeto(
        id=p.id,
        nome=p.nome_projeto,
        data_aceite=p.data_aceite,
        data_conclusao=p.data_conclusao,
        status=p.status,
        valor_mensalidades=p.valor_mensalidades or 0,
        responsavel=responsavel,
        descricao=p.descricao or '',
        numero_unidades=p.numero_unidades or 1,
        potencial_cliente=p.potencial_cliente or 'Médio',
        tipo_projeto=p.tipo_projeto or 'Novo'
    )

    # Adiciona módulos
    modulos_db = ERPModuloDB.query.filter_by(projeto_id=p.id).all()
    for m in modulos_db:
        mod = Modulo(
            id=m.id,
            projeto_id=m.projeto_id,
            nome=m.modulo,
            status=m.status_modulo,
            data_inicio=m.data_inicio_modulo,
            data_conclusao=m.data_conclusao_modulo,
            percentual_conclusao=m.percentual_conclusao or 0
        )
        proj.adicionar_modulo(mod)

    # Adiciona unidades
    unidades_db = ERPUnidadeDB.query.filter_by(projeto_id=p.id).all()
    for u in unidades_db:
        uni = Unidade(
            id=u.id,
            projeto_id=u.projeto_id,
            nome=u.unidade,
            status=u.status_unidade,
            data_inicio=u.data_inicio_unidade,
            data_conclusao=u.data_conclusao_unidade
        )
        proj.adicionar_unidade(uni)

    # Adiciona atividades
    atividades_db = ERPAtividadeDB.query.filter_by(projeto_id=p.id).order_by(ERPAtividadeDB.data_reuniao.desc()).all()
    for a in atividades_db:
        ativ = Atividade(
            id=a.id,
            projeto_id=a.projeto_id,
            titulo=a.titulo,
            data_reuniao=a.data_reuniao,
            descricao=a.descricao or '',
            responsavel_nota=a.responsavel_nota or '',
            status=a.status_atividade,
            concluida=a.concluida or False
        )
        proj.adicionar_atividade(ativ)

    return proj

def carregar_tudo():
    colabs_db    = ColaboradorDB.query.filter_by(ativo=True).all()
    colaboradores = [db_to_colab(c) for c in colabs_db]
    ferias_todos = FeriasDB.query.filter(FeriasDB.status != 'Cancelado').all()
    ferias_plan  = [db_to_ferias(f) for f in ferias_todos if f.status != 'Realizado']
    ferias_real  = [db_to_ferias(f) for f in ferias_todos if f.status == 'Realizado']
    return colaboradores, ferias_plan, ferias_real

def obter_coordenadores():
    """Retorna lista de colaboradores com cargo de coordenador ou gerencia.
    Se nenhum for encontrado com esses cargos, retorna todos os ativos."""
    colabs_db = ColaboradorDB.query.filter_by(ativo=True).order_by(ColaboradorDB.nome).all()
    coordenadores = []
    for c in colabs_db:
        cargo, _ = parse_cargo_uf(c.time)
        if cargo in ('coordenador', 'gerencia'):
            coordenadores.append(c)
    # Se nenhum coordenador/gerente cadastrado, exibe todos os colaboradores ativos
    if not coordenadores:
        return colabs_db
    return coordenadores

# ─── Helpers de saldo / datas ─────────────────────────────────────────────────

def saldo_colab(colab, ferias_real, ferias_plan=None):
    cargo, _ = parse_cargo_uf(colab.time)
    if cargo == 'estagiario':
        return None
    base = colab.calcular_saldo_ferias(ferias_real)
    if ferias_plan:
        dias_p = sum(
            f.dias for f in ferias_plan
            if f.colaborador_id == colab.id and f.status != 'Cancelado'
        )
        base -= dias_p
    return max(0, base)

def status_saldo(saldo):
    if saldo is None: return 'estagiario'
    if saldo <= 0:    return 'critico'
    if saldo <= 10:   return 'alerta'
    return 'ok'

def saldo_quantizado(saldo_raw):
    if saldo_raw is None:
        return None
    return min(30, (saldo_raw // 10) * 10)

def proximo_dia_util(dt):
    wd = dt.weekday()
    if wd == 5: return dt + timedelta(days=2)
    if wd == 6: return dt + timedelta(days=1)
    return dt

def calcular_data_limite(colab, saldo_raw):
    if saldo_raw is None or saldo_raw <= 0:
        return None
    hoje = date.today()
    adm = colab.data_admissao.date() if isinstance(colab.data_admissao, datetime) else colab.data_admissao
    try:
        proximo_aniv = adm.replace(year=hoje.year)
    except ValueError:
        proximo_aniv = adm.replace(year=hoje.year, day=28)
    if proximo_aniv <= hoje:
        try:
            proximo_aniv = adm.replace(year=hoje.year + 1)
        except ValueError:
            proximo_aniv = adm.replace(year=hoje.year + 1, day=28)
    return proximo_aniv - timedelta(days=saldo_raw)

def tempo_casa_str(data_admissao):
    delta = relativedelta(datetime.now(), data_admissao)
    partes = []
    if delta.years:  partes.append(f"{delta.years}a")
    if delta.months: partes.append(f"{delta.months}m")
    return ' '.join(partes) if partes else '< 1 mês'

def build_calendario(ano, mes, ferias_plan, colaboradores):
    colab_map = {c.id: c for c in colaboradores}
    hoje = datetime.now().date()
    primeiro_dia = date(ano, mes, 1)
    _, n_dias = calendar.monthrange(ano, mes)
    inicio_semana = (primeiro_dia.weekday() + 1) % 7
    grade, semana = [], [0] * inicio_semana
    for dia in range(1, n_dias + 1):
        semana.append(dia)
        if len(semana) == 7:
            grade.append(semana); semana = []
    if semana:
        semana += [0] * (7 - len(semana)); grade.append(semana)
    cal_info = {}
    for dia in range(1, n_dias + 1):
        d = date(ano, mes, dia)
        em_ferias = []
        for f in ferias_plan:
            if f.status in ('Cancelado', 'Solicitado'): continue
            if f.data_inicio.date() <= d <= f.data_fim.date():
                colab = colab_map.get(f.colaborador_id)
                if colab: em_ferias.append(colab)
        fds = d.weekday() >= 5
        cls_parts = ['cal-day']
        if d == hoje: cls_parts.append('today')
        if fds: cls_parts.append('fds')
        tip = ''
        if len(em_ferias) > 1:
            cls_parts.append('ferias-multi')
            tip = ', '.join(c.nome for c in em_ferias)
        elif len(em_ferias) == 1:
            idx = next((i for i, c in enumerate(colaboradores) if c.id == em_ferias[0].id), 0)
            cls_parts.append(f'ferias-{(idx % 5) + 1}')
            tip = em_ferias[0].nome
        cal_info[dia] = {'cls': ' '.join(cls_parts), 'tip': tip, 'em_ferias': em_ferias}
    return grade, cal_info

def build_legenda(colaboradores, ferias_plan):
    ids_com_ferias = {f.colaborador_id for f in ferias_plan
                      if f.status not in ('Cancelado', 'Solicitado')}
    legenda, bg = [], ['#bbdefb', '#c8e6c9', '#f8bbd0', '#ffe0b2', '#e1bee7']
    idx = 0
    for c in colaboradores:
        if c.id in ids_com_ferias:
            legenda.append({'nome': c.nome.split()[0], 'bg': bg[idx % 5]})
            idx += 1
            if idx >= 5: break
    return legenda

# ─── Rotas: autenticação ──────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        senha    = request.form.get('senha', '')
        user = User.query.filter_by(username=username, ativo=True).first()
        if user and user.check_senha(senha):
            login_user(user, remember=True)
            return redirect(request.args.get('next') or url_for('index'))
        flash('Usuário ou senha incorretos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ─── Rotas: gestão de usuários ────────────────────────────────────────────────

@app.route('/usuarios')
@login_required
@gestor_required
def listar_usuarios():
    usuarios     = User.query.order_by(User.nome).all()
    colaboradores = ColaboradorDB.query.filter_by(ativo=True).order_by(ColaboradorDB.nome).all()
    colab_map    = {c.id: c for c in colaboradores}
    return render_template('usuarios.html', usuarios=usuarios, colab_map=colab_map)

@app.route('/novo-usuario', methods=['GET', 'POST'])
@login_required
@gestor_required
def novo_usuario():
    colaboradores = ColaboradorDB.query.filter_by(ativo=True).order_by(ColaboradorDB.nome).all()
    if request.method == 'POST':
        nome     = request.form.get('nome', '').strip()
        username = request.form.get('username', '').strip().lower()
        senha    = request.form.get('senha', '')
        senha2   = request.form.get('senha2', '')
        perfil   = request.form.get('perfil', 'colaborador')
        colab_id = request.form.get('colaborador_id') or None
        if colab_id: colab_id = int(colab_id)

        if not nome or not username or not senha:
            return render_template('novo_usuario.html', colaboradores=colaboradores,
                                   erro='Preencha todos os campos obrigatórios.')
        if senha != senha2:
            return render_template('novo_usuario.html', colaboradores=colaboradores,
                                   erro='As senhas não coincidem.')
        if User.query.filter_by(username=username).first():
            return render_template('novo_usuario.html', colaboradores=colaboradores,
                                   erro=f'Username "{username}" já está em uso.')

        u = User(nome=nome, username=username, perfil=perfil, colaborador_id=colab_id)
        u.set_senha(senha)
        db.session.add(u)
        db.session.commit()
        flash(f'Usuário "{nome}" criado com sucesso!', 'success')
        return redirect(url_for('listar_usuarios'))

    return render_template('novo_usuario.html', colaboradores=colaboradores)

@app.route('/editar-usuario/<int:uid>', methods=['GET', 'POST'])
@login_required
@gestor_required
def editar_usuario(uid):
    user = User.query.get_or_404(uid)
    colaboradores = ColaboradorDB.query.filter_by(ativo=True).order_by(ColaboradorDB.nome).all()
    if request.method == 'POST':
        nome     = request.form.get('nome', '').strip()
        perfil   = request.form.get('perfil', user.perfil)
        colab_id = request.form.get('colaborador_id') or None
        if colab_id: colab_id = int(colab_id)
        senha    = request.form.get('senha', '')
        senha2   = request.form.get('senha2', '')
        ativo    = request.form.get('ativo') == '1'

        if not nome:
            return render_template('novo_usuario.html', user=user, colaboradores=colaboradores,
                                   erro='Nome é obrigatório.')
        if senha and senha != senha2:
            return render_template('novo_usuario.html', user=user, colaboradores=colaboradores,
                                   erro='As senhas não coincidem.')

        user.nome           = nome
        user.perfil         = perfil
        user.colaborador_id = colab_id
        user.ativo          = ativo
        if senha:
            user.set_senha(senha)
        db.session.commit()
        flash(f'Usuário "{nome}" atualizado.', 'success')
        return redirect(url_for('listar_usuarios'))

    return render_template('novo_usuario.html', user=user, colaboradores=colaboradores)

@app.route('/excluir-usuario/<int:uid>', methods=['POST'])
@login_required
@gestor_required
def excluir_usuario(uid):
    if uid == current_user.id:
        flash('Você não pode excluir seu próprio usuário.', 'danger')
        return redirect(url_for('listar_usuarios'))
    user = User.query.get_or_404(uid)
    user.ativo = False
    db.session.commit()
    flash(f'Usuário "{user.nome}" desativado.', 'warning')
    return redirect(url_for('listar_usuarios'))

# ─── Rota: Dashboard ──────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    colaboradores_raw, ferias_plan, ferias_real = carregar_tudo()
    colaboradores = sort_colaboradores(colaboradores_raw)
    analytics     = FeriasAnalytics(colaboradores, ferias_plan, ferias_real)
    colab_map     = {c.id: c for c in colaboradores}
    hoje          = datetime.now()
    ano  = int(request.args.get('ano', hoje.year))
    mes  = int(request.args.get('mes', hoje.month))

    total_ativos = sum(1 for c in colaboradores if c.ativo)

    contagem_cargos = {'gerencia': 0, 'coordenador': 0, 'consultor': 0, 'estagiario': 0}
    for c in colaboradores:
        if c.ativo:
            cargo, _ = parse_cargo_uf(c.time)
            contagem_cargos[cargo] = contagem_cargos.get(cargo, 0) + 1

    em_ferias_hoje = sum(
        1 for f in ferias_plan
        if f.status not in ('Cancelado', 'Solicitado')
        and f.data_inicio.date() <= hoje.date() <= f.data_fim.date()
    )
    prox30 = hoje + timedelta(days=30)
    proximos_30 = sum(
        1 for f in ferias_plan
        if f.status not in ('Cancelado', 'Solicitado')
        and hoje.date() <= f.data_inicio.date() <= prox30.date()
    )

    conflitos = analytics.detectar_conflitos()
    conflitos_pendentes = [
        c for c in conflitos
        if not c['aprovado'] and c['ferias_principal'].status != 'Solicitado'
    ]
    total_conflitos = len(conflitos_pendentes)

    # Solicitações pendentes de aprovação (gestor)
    solicitacoes_pendentes = []
    if current_user.is_gestor:
        for s_db in FeriasDB.query.filter_by(status='Solicitado').all():
            colab = colab_map.get(s_db.colaborador_id)
            if colab:
                idx = next((i for i, c in enumerate(colaboradores) if c.id == colab.id), 0)
                solicitacoes_pendentes.append({
                    'ferias':      db_to_ferias(s_db),
                    'colaborador': colab,
                    'cor':         cor_colab(idx),
                })

    # Próximas férias (90 dias) — exclui Solicitado
    proximas_raw = analytics.obter_proximas_ferias(90)
    proximas = []
    for item in proximas_raw:
        if item['ferias'].status == 'Solicitado': continue
        c = item['colaborador']
        if c:
            idx = next((i for i, x in enumerate(colaboradores) if x.id == c.id), 0)
            proximas.append({**item, 'cor': cor_colab(idx)})

    # Saldos
    saldos_lista = []
    for i, c in enumerate(colaboradores):
        if not c.ativo: continue
        cargo, _ = parse_cargo_uf(c.time)
        if cargo == 'estagiario': continue
        s_raw  = saldo_colab(c, ferias_real, ferias_plan)
        s_disp = saldo_quantizado(s_raw)
        saldos_lista.append({
            'colaborador': c, 'saldo': s_disp, 'saldo_raw': s_raw,
            'status': status_saldo(s_raw), 'cor': cor_colab(i),
        })
    saldos_lista.sort(key=lambda x: (x['saldo_raw'] is None, x['saldo_raw'] or 0))

    # Minhas férias (colaborador)
    minhas_ferias = []
    if not current_user.is_gestor and current_user.colaborador_id:
        minhas_ferias = sorted(
            [f for f in ferias_plan if f.colaborador_id == current_user.colaborador_id],
            key=lambda f: f.data_inicio,
        )

    calendario, cal_info = build_calendario(ano, mes, ferias_plan, colaboradores)
    mes_anterior = date(ano, mes, 1) - timedelta(days=1)
    mes_proximo  = (date(ano, mes, 1) + timedelta(days=32)).replace(day=1)
    meses_pt = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    legenda_cal = build_legenda(colaboradores, ferias_plan)

    # Dados de projetos ERP
    projetos_db = ERPProjetoDB.query.all()
    projetos = [db_to_projeto(p) for p in projetos_db]
    proj_analytics = ProjetoAnalytics(projetos)
    proj_resumo = proj_analytics.resumo_geral()
    proj_criticos = proj_analytics.projetos_criticos()

    return render_template('dashboard.html',
        hoje=hoje, ano=ano, mes=mes,
        mes_nome=meses_pt[mes], mes_anterior=mes_anterior, mes_proximo=mes_proximo,
        total_ativos=total_ativos, em_ferias_hoje=em_ferias_hoje,
        proximos_30=proximos_30, total_conflitos=total_conflitos,
        conflitos_pendentes=conflitos_pendentes, colaboradores_dict=colab_map,
        proximas=proximas, saldos_lista=saldos_lista,
        calendario=calendario, cal_info=cal_info, legenda_cal=legenda_cal,
        contagem_cargos=contagem_cargos,
        solicitacoes_pendentes=solicitacoes_pendentes,
        minhas_ferias=minhas_ferias,
        proj_resumo=proj_resumo,
        proj_criticos=proj_criticos,
        projetos=projetos,
    )

# ─── Rota: Timeline ───────────────────────────────────────────────────────────

@app.route('/timeline')
@login_required
def timeline():
    colaboradores_raw, ferias_plan, ferias_real = carregar_tudo()
    colaboradores = sort_colaboradores(colaboradores_raw)
    hoje = datetime.now()

    ini_str = request.args.get('inicio', hoje.strftime('%Y-%m'))
    fim_str = request.args.get('fim', (hoje + relativedelta(months=5)).strftime('%Y-%m'))
    try:
        ini_date = datetime.strptime(ini_str + '-01', '%Y-%m-%d').date()
        fim_date = (datetime.strptime(fim_str + '-01', '%Y-%m-%d') + relativedelta(months=1) - timedelta(days=1)).date()
    except Exception:
        ini_date = hoje.date().replace(day=1)
        fim_date = (hoje + relativedelta(months=5)).date()

    total_dias = (fim_date - ini_date).days + 1
    meses_pt   = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                  'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    meses_header = []
    cur = ini_date.replace(day=1)
    while cur <= fim_date:
        _, nd = calendar.monthrange(cur.year, cur.month)
        m_ini = max(cur, ini_date)
        m_fim = min(date(cur.year, cur.month, nd), fim_date)
        dias_mes = (m_fim - m_ini).days + 1
        meses_header.append({'label': f"{meses_pt[cur.month]}/{str(cur.year)[2:]}", 'dias': dias_mes})
        cur = (cur + relativedelta(months=1)).replace(day=1)

    hoje_pct = None
    if ini_date <= hoje.date() <= fim_date:
        hoje_pct = round(((hoje.date() - ini_date).days / total_dias) * 100, 2)

    from collections import defaultdict
    grupos_raw = defaultdict(list)
    for i, colab in enumerate(colaboradores):
        if not colab.ativo: continue
        cargo_tl, uf_tl = parse_cargo_uf(colab.time)
        if cargo_tl == 'estagiario': continue
        segs = []
        for f in ferias_plan:
            if f.colaborador_id != colab.id or f.status in ('Cancelado', 'Solicitado'): continue
            f_ini = f.data_inicio.date(); f_fim = f.data_fim.date()
            if f_fim < ini_date or f_ini > fim_date: continue
            clip_ini = max(f_ini, ini_date); clip_fim = min(f_fim, fim_date)
            left  = round(((clip_ini - ini_date).days / total_dias) * 100, 2)
            width = round(((clip_fim - clip_ini).days + 1) / total_dias * 100, 2)
            opacity = '1' if f.status == 'Confirmado' else '0.55'
            label = f"{colab.nome}: {f_ini.strftime('%d/%m')}–{f_fim.strftime('%d/%m')} ({f.dias}d) · {f.status}"
            segs.append({'left': left, 'width': width, 'opacity': opacity,
                         'dias': f.dias, 'label': label, 'fid': f.id,
                         'data_inicio_real': f_ini.strftime('%Y-%m-%d'),
                         'data_fim_real':    f_fim.strftime('%Y-%m-%d'),
                         'status': f.status})
        grupos_raw[cargo_tl].append({
            'nome': colab.nome, 'uf': uf_tl, 'cor': cor_colab(i),
            'segmentos': segs, 'tem_ferias': bool(segs),
            'id': colab.id, 'cargo': cargo_tl,
        })

    dados_timeline = []
    for cargo_key in sorted(grupos_raw.keys(), key=lambda k: CARGO_ORDEM.get(k, 9)):
        membros = sorted(grupos_raw[cargo_key], key=lambda x: (x['uf'], x['nome'].lower()))
        dados_timeline.append({
            'separador': True, 'tipo': 'cargo',
            'label': CARGO_LABEL.get(cargo_key, cargo_key.title()),
            'total': len(membros), 'cargo_key': cargo_key,
        })
        prev_uf = None
        for m in membros:
            cur_uf = m['uf'] or ''
            if cur_uf != prev_uf:
                prev_uf = cur_uf
                dados_timeline.append({
                    'separador': True, 'tipo': 'uf',
                    'label': cur_uf or 'UF não informada',
                    'cargo_key': cargo_key, 'uf_key': cur_uf,
                })
            dados_timeline.append({'separador': False, **m})

    meses_pt_full = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                     'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    periodo_label = (f"{meses_pt_full[ini_date.month]}/{ini_date.year} → "
                     f"{meses_pt_full[fim_date.month]}/{fim_date.year}")

    return render_template('timeline.html',
        filtro_inicio=ini_str, filtro_fim=fim_str,
        periodo_label=periodo_label, meses_header=meses_header,
        dados_timeline=dados_timeline, hoje_pct=hoje_pct,
    )

# ─── Rota: Colaboradores ──────────────────────────────────────────────────────

@app.route('/colaboradores')
@login_required
def listar_colaboradores():
    colaboradores_raw, ferias_plan, ferias_real = carregar_tudo()
    colaboradores = [c for c in sort_colaboradores(colaboradores_raw) if c.ativo]
    hoje   = datetime.now()
    filtro = request.args.get('filtro', '')

    ids_filtro = None
    if filtro == 'ferias_hoje':
        ids_filtro = {f.colaborador_id for f in ferias_plan
                      if f.status != 'Cancelado'
                      and f.data_inicio.date() <= hoje.date() <= f.data_fim.date()}
    elif filtro == 'proximos30':
        prox30 = hoje + timedelta(days=30)
        ids_filtro = {f.colaborador_id for f in ferias_plan
                      if f.status != 'Cancelado'
                      and hoje.date() <= f.data_inicio.date() <= prox30.date()}

    from collections import defaultdict
    itens_map = {}
    for i, c in enumerate(colaboradores):
        if ids_filtro is not None and c.id not in ids_filtro: continue
        cargo, uf = parse_cargo_uf(c.time)
        s_raw  = saldo_colab(c, ferias_real, ferias_plan)
        s_disp = saldo_quantizado(s_raw)
        proxima = next(
            (f for f in sorted(ferias_plan, key=lambda x: x.data_inicio)
             if f.colaborador_id == c.id and f.status not in ('Cancelado', 'Solicitado')
             and f.data_inicio >= hoje),
            None,
        )
        data_lim = calcular_data_limite(c, s_raw) if cargo != 'estagiario' else None
        itens_map[c.id] = {
            'colaborador': c, 'saldo': s_disp, 'saldo_raw': s_raw,
            'status_saldo': status_saldo(s_raw), 'cor': cor_colab(i),
            'tempo_casa': tempo_casa_str(c.data_admissao), 'proxima_ferias': proxima,
            'cargo': cargo, 'uf': uf, 'data_limite': data_lim,
        }

    grupos_raw = defaultdict(list)
    for item in itens_map.values():
        grupos_raw[item['cargo']].append(item)

    grupos = []
    for cargo_key in sorted(grupos_raw.keys(), key=lambda k: CARGO_ORDEM.get(k, 9)):
        membros = sorted(grupos_raw[cargo_key], key=lambda x: (x['uf'], x['colaborador'].nome.lower()))
        grupos.append({'label': CARGO_LABEL.get(cargo_key, cargo_key.title()),
                       'key': cargo_key, 'itens': membros})

    filtro_label = {'ferias_hoje': 'Em férias hoje', 'proximos30': 'Saem em 30 dias'}.get(filtro, '')
    return render_template('colaboradores.html', grupos=grupos, filtro_ativo=filtro,
                           filtro_label=filtro_label, hoje=hoje)

# ─── Rota: Novo colaborador ───────────────────────────────────────────────────

@app.route('/novo-colaborador', methods=['GET', 'POST'])
@login_required
@gestor_required
def novo_colaborador():
    if request.method == 'POST':
        nome      = request.form.get('nome', '').strip()
        data_str  = request.form.get('data_admissao', '')
        tipo      = request.form.get('tipo', 'Consultor').strip()
        uf_form   = request.form.get('uf', '').strip().upper()
        cidade    = request.form.get('cidade', '').strip()
        time_full = f"{tipo} - {uf_form}" if uf_form else tipo
        try:
            data_admissao = datetime.strptime(data_str, '%Y-%m-%d')
        except ValueError:
            return render_template('novo_colaborador.html', erro='Data de admissão inválida.',
                                   form_nome=nome, form_uf=uf_form, form_cidade=cidade)
        valido, erro = FeriasValidator.validar_novo_colaborador(nome, data_admissao)
        if not valido:
            return render_template('novo_colaborador.html', erro=erro,
                                   form_nome=nome, form_data=data_str,
                                   form_uf=uf_form, form_cidade=cidade)
        novo = ColaboradorDB(nome=nome, data_admissao=data_admissao.date(),
                             time=time_full, cidade=cidade)
        db.session.add(novo)
        db.session.commit()
        flash(f'Colaborador "{nome}" adicionado com sucesso!', 'success')
        return redirect(url_for('listar_colaboradores'))
    return render_template('novo_colaborador.html')

# ─── Rota: Editar colaborador ─────────────────────────────────────────────────

@app.route('/editar-colaborador/<int:cid>', methods=['GET', 'POST'])
@login_required
@gestor_required
def editar_colaborador(cid):
    c_db = ColaboradorDB.query.get_or_404(cid)
    cargo_atual, uf_atual = parse_cargo_uf(c_db.time)
    if request.method == 'POST':
        nome      = request.form.get('nome', '').strip()
        data_str  = request.form.get('data_admissao', '')
        tipo      = request.form.get('tipo', 'Consultor').strip()
        uf_form   = request.form.get('uf', '').strip().upper()
        cidade    = request.form.get('cidade', '').strip()
        time_full = f"{tipo} - {uf_form}" if uf_form else tipo
        try:
            data_admissao = datetime.strptime(data_str, '%Y-%m-%d')
        except ValueError:
            colab = db_to_colab(c_db)
            return render_template('editar_colaborador.html', colab=colab,
                                   cargo_atual=cargo_atual, uf_atual=uf_atual,
                                   erro='Data de admissão inválida.')
        valido, erro = FeriasValidator.validar_novo_colaborador(nome, data_admissao)
        if not valido:
            colab = db_to_colab(c_db)
            return render_template('editar_colaborador.html', colab=colab,
                                   cargo_atual=cargo_atual, uf_atual=uf_atual, erro=erro)
        c_db.nome          = nome
        c_db.data_admissao = data_admissao.date()
        c_db.time          = time_full
        c_db.cidade        = cidade
        db.session.commit()
        flash(f'Dados de "{nome}" atualizados!', 'success')
        return redirect(url_for('listar_colaboradores'))
    colab = db_to_colab(c_db)
    return render_template('editar_colaborador.html', colab=colab,
                           cargo_atual=cargo_atual, uf_atual=uf_atual)

# ─── Rota: Excluir colaborador ────────────────────────────────────────────────

@app.route('/excluir-colaborador/<int:cid>', methods=['POST'])
@login_required
@gestor_required
def excluir_colaborador(cid):
    c_db = ColaboradorDB.query.get_or_404(cid)
    c_db.ativo = False
    FeriasDB.query.filter_by(colaborador_id=cid).filter(
        FeriasDB.status.notin_(['Cancelado', 'Realizado'])
    ).update({'status': 'Cancelado'})
    db.session.commit()
    flash(f'"{c_db.nome}" removido da equipe.', 'warning')
    return redirect(url_for('listar_colaboradores'))

# ─── Rota: Nova / Solicitar férias ────────────────────────────────────────────

@app.route('/nova-ferias', methods=['GET', 'POST'])
@login_required
def nova_ferias():
    colaboradores_raw, ferias_plan, ferias_real = carregar_tudo()
    colaboradores = [c for c in colaboradores_raw if parse_cargo_uf(c.time)[0] != 'estagiario']

    modo_solicitacao = not current_user.is_gestor
    if modo_solicitacao:
        if not current_user.colaborador_id:
            flash('Seu usuário não está vinculado a um colaborador. Fale com o gestor.', 'warning')
            return redirect(url_for('index'))
        sel_colab    = current_user.colaborador_id
        colaboradores = [c for c in colaboradores if c.id == sel_colab]
    else:
        sel_colab = request.args.get('colab', type=int)

    if request.method == 'POST':
        if modo_solicitacao:
            colab_id   = current_user.colaborador_id
            status_sel = 'Solicitado'
        else:
            colab_id   = int(request.form.get('colaborador_id', 0))
            status_sel = request.form.get('status', 'Planejado')

        ini_str   = request.form.get('data_inicio', '')
        fim_str   = request.form.get('data_fim', '')
        confirmar = request.form.get('confirmar', '')

        colab = next((c for c in colaboradores_raw if c.id == colab_id), None)
        if not colab:
            return render_template('nova_ferias.html', colaboradores=colaboradores,
                                   erro='Colaborador não encontrado.',
                                   sel_colab=sel_colab,
                                   modo_solicitacao=modo_solicitacao)
        try:
            ini = datetime.strptime(ini_str, '%Y-%m-%d')
            fim = datetime.strptime(fim_str, '%Y-%m-%d')
        except ValueError:
            return render_template('nova_ferias.html', colaboradores=colaboradores,
                                   erro='Datas inválidas.',
                                   sel_colab=colab_id, form_ini=ini_str, form_fim=fim_str,
                                   modo_solicitacao=modo_solicitacao)

        saldo    = saldo_colab(colab, ferias_real, ferias_plan)
        nova_f   = Ferias(0, colab_id, ini, fim, status_sel)

        # Colaborador só precisa validar datas (gestor cuida do restante ao aprovar)
        if not modo_solicitacao:
            valido, erro = FeriasValidator.validar_ferias(nova_f, colab, ferias_plan, ferias_real, saldo or 0)
            if not valido:
                return render_template('nova_ferias.html', colaboradores=colaboradores,
                                       erro=erro, sel_colab=colab_id,
                                       form_ini=ini_str, form_fim=fim_str,
                                       modo_solicitacao=modo_solicitacao)

            cor_map  = {c.id: cor_colab(i) for i, c in enumerate(colaboradores_raw)}
            conflitos = FeriasValidator.detectar_conflitos_time(
                nova_f, ferias_plan, {c.id: c for c in colaboradores_raw})
            if conflitos and not confirmar:
                conflitos_info = []
                for fc, dias in conflitos:
                    cc  = next((c for c in colaboradores_raw if c.id == fc.colaborador_id), None)
                    idx = next((i for i, c in enumerate(colaboradores_raw) if c.id == fc.colaborador_id), 0)
                    conflitos_info.append({'ferias': fc, 'colaborador': cc,
                                           'dias_overlap': dias, 'cor': cor_colab(idx)})
                idx_colab = next((i for i, c in enumerate(colaboradores_raw) if c.id == colab_id), 0)
                return render_template('confirmar_conflito.html',
                    colaborador=colab, nova_ferias=nova_f,
                    conflitos=conflitos_info, cor_colab=cor_colab(idx_colab))
        else:
            # Validação básica de datas para solicitações
            if fim <= ini:
                return render_template('nova_ferias.html', colaboradores=colaboradores,
                                       erro='Data final deve ser posterior à data inicial.',
                                       sel_colab=colab_id, form_ini=ini_str, form_fim=fim_str,
                                       modo_solicitacao=modo_solicitacao)
            conflitos = []
            confirmar = False

        f_db = FeriasDB(
            colaborador_id=colab_id,
            data_inicio=ini.date(), data_fim=fim.date(), dias=nova_f.dias,
            status=status_sel,
            conflito_detectado=bool(conflitos),
            conflito_aprovado=bool(conflitos and confirmar),
        )
        db.session.add(f_db)
        db.session.commit()

        if modo_solicitacao:
            flash('Solicitação enviada! O gestor irá analisar em breve.', 'success')
        else:
            flash(f'Férias de {colab.nome} registradas com sucesso!', 'success')
        return redirect(url_for('index'))

    return render_template('nova_ferias.html', colaboradores=colaboradores,
                           sel_colab=sel_colab, modo_solicitacao=modo_solicitacao)

# ─── Rota: Aprovar / Rejeitar solicitações ────────────────────────────────────

@app.route('/aprovar-solicitacao/<int:fid>', methods=['POST'])
@login_required
@gestor_required
def aprovar_solicitacao(fid):
    f_db = FeriasDB.query.get_or_404(fid)
    f_db.status = 'Planejado'
    db.session.commit()
    colab = ColaboradorDB.query.get(f_db.colaborador_id)
    flash(f'Solicitação de {colab.nome if colab else "colaborador"} aprovada!', 'success')
    return redirect(url_for('index'))

@app.route('/rejeitar-solicitacao/<int:fid>', methods=['POST'])
@login_required
@gestor_required
def rejeitar_solicitacao(fid):
    f_db = FeriasDB.query.get_or_404(fid)
    f_db.status = 'Cancelado'
    db.session.commit()
    colab = ColaboradorDB.query.get(f_db.colaborador_id)
    flash(f'Solicitação de {colab.nome if colab else "colaborador"} rejeitada.', 'warning')
    return redirect(url_for('index'))

# ─── Rotas: operações em férias ───────────────────────────────────────────────

@app.route('/deletar-ferias/<int:fid>', methods=['POST'])
@login_required
@gestor_required
def deletar_ferias(fid):
    f_db = FeriasDB.query.get_or_404(fid)
    f_db.status = 'Cancelado'
    db.session.commit()
    flash('Férias canceladas.', 'warning')
    next_url = request.form.get('next', url_for('index'))
    return redirect(next_url)

@app.route('/editar-ferias/<int:fid>', methods=['POST'])
@login_required
@gestor_required
def editar_ferias(fid):
    f_db = FeriasDB.query.get_or_404(fid)
    ini_str    = request.form.get('data_inicio', '')
    fim_str    = request.form.get('data_fim', '')
    status_sel = request.form.get('status', f_db.status)
    try:
        ini = datetime.strptime(ini_str, '%Y-%m-%d')
        fim = datetime.strptime(fim_str, '%Y-%m-%d')
    except ValueError:
        flash('Datas inválidas.', 'danger')
        return redirect(url_for('timeline'))
    colaboradores_raw, ferias_plan, ferias_real = carregar_tudo()
    colab = next((c for c in colaboradores_raw if c.id == f_db.colaborador_id), None)
    if not colab:
        flash('Colaborador não encontrado.', 'danger')
        return redirect(url_for('timeline'))
    outras    = [x for x in ferias_plan if x.id != fid and x.status != 'Cancelado']
    nova_f    = Ferias(fid, f_db.colaborador_id, ini, fim, status_sel)
    saldo_base = saldo_colab(colab, ferias_real, [x for x in ferias_plan if x.id != fid])
    valido, erro = FeriasValidator.validar_ferias(nova_f, colab, outras, ferias_real, saldo_base or 0)
    if not valido:
        flash(f'Erro: {erro}', 'danger')
        return redirect(url_for('timeline'))
    f_db.data_inicio = ini.date()
    f_db.data_fim    = fim.date()
    f_db.dias        = nova_f.dias
    f_db.status      = status_sel
    db.session.commit()
    flash(f'Férias de {colab.nome} atualizadas!', 'success')
    return redirect(url_for('timeline'))

@app.route('/confirmar-ferias/<int:fid>', methods=['POST'])
@login_required
@gestor_required
def confirmar_ferias(fid):
    f_db = FeriasDB.query.get_or_404(fid)
    f_db.status = 'Confirmado'
    db.session.commit()
    colab = ColaboradorDB.query.get(f_db.colaborador_id)
    flash(f'Férias de {colab.nome if colab else "colaborador"} confirmadas!', 'success')
    return redirect(url_for('index'))

@app.route('/reprovar-ferias/<int:fid>', methods=['POST'])
@login_required
@gestor_required
def reprovar_ferias(fid):
    f_db = FeriasDB.query.get_or_404(fid)
    f_db.status = 'Cancelado'
    db.session.commit()
    colab = ColaboradorDB.query.get(f_db.colaborador_id)
    flash(f'Férias de {colab.nome if colab else "colaborador"} canceladas.', 'danger')
    return redirect(url_for('index'))

@app.route('/aprovar-conflito/<int:fid>', methods=['POST'])
@login_required
@gestor_required
def aprovar_conflito(fid):
    f_db = FeriasDB.query.get_or_404(fid)
    f_db.conflito_aprovado = True
    db.session.commit()
    flash('Conflito aprovado.', 'success')
    return redirect(url_for('index'))

# ─── Rotas: comissionamento manual ─────────────────────────────────────────────

@app.route('/comissionamentos')
@login_required
def listar_comissionamentos():
    """Lista comissionamentos manuais"""
    comissions_db = ComissionamentoDB.query.order_by(ComissionamentoDB.data_comissao.desc()).all()

    return render_template('comissionamentos/lista_comissionamentos.html',
                          comissionamentos=comissions_db)

@app.route('/novo-comissionamento', methods=['GET', 'POST'])
@login_required
@gestor_required
def novo_comissionamento():
    """Cria novo comissionamento manual"""
    colaboradores = ColaboradorDB.query.filter_by(ativo=True).order_by(ColaboradorDB.nome).all()

    if request.method == 'POST':
        consultor_id = request.form.get('consultor_id')
        cliente = request.form.get('cliente', '').strip()
        data_str = request.form.get('data_comissao', '')
        horas_str = request.form.get('horas_comissionadas', '0')
        hora_fora_estado = request.form.get('hora_fora_estado', '').strip()
        motivo = request.form.get('motivo', '').strip()

        # Validações básicas
        if not cliente:
            flash('Cliente é obrigatório.', 'danger')
            return render_template('comissionamentos/novo_comissionamento.html',
                                   colaboradores=colaboradores)

        try:
            data_comissao = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Data inválida.', 'danger')
            return render_template('comissionamentos/novo_comissionamento.html',
                                   colaboradores=colaboradores)

        try:
            horas = float(horas_str.replace(',', '.')) if horas_str else 0
        except ValueError:
            horas = 0

        novo = ComissionamentoDB(
            consultor_id=int(consultor_id) if consultor_id else None,
            cliente=cliente,
            data_comissao=data_comissao,
            horas_comissionadas=horas,
            hora_fora_estado=hora_fora_estado,
            motivo=motivo
        )
        db.session.add(novo)
        db.session.commit()

        flash(f'Comissionamento "{cliente}" adicionado!', 'success')
        return redirect(url_for('listar_comissionamentos'))

    return render_template('comissionamentos/novo_comissionamento.html',
                          colaboradores=colaboradores)

@app.route('/editar-comissionamento/<int:cid>', methods=['GET', 'POST'])
@login_required
@gestor_required
def editar_comissionamento(cid):
    """Edita comissionamento manual"""
    comissao_db = ComissionamentoDB.query.get_or_404(cid)
    colaboradores = ColaboradorDB.query.filter_by(ativo=True).order_by(ColaboradorDB.nome).all()

    if request.method == 'POST':
        consultor_id = request.form.get('consultor_id')
        cliente = request.form.get('cliente', '').strip()
        data_str = request.form.get('data_comissao', '')
        horas_str = request.form.get('horas_comissionadas', '0')
        hora_fora_estado = request.form.get('hora_fora_estado', '').strip()
        motivo = request.form.get('motivo', '').strip()

        if not cliente:
            flash('Cliente é obrigatório.', 'danger')
            return redirect(url_for('editar_comissionamento', cid=cid))

        try:
            data_comissao = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Data inválida.', 'danger')
            return redirect(url_for('editar_comissionamento', cid=cid))

        try:
            horas = float(horas_str.replace(',', '.')) if horas_str else 0
        except ValueError:
            horas = 0

        comissao_db.consultor_id = int(consultor_id) if consultor_id else None
        comissao_db.cliente = cliente
        comissao_db.data_comissao = data_comissao
        comissao_db.horas_comissionadas = horas
        comissao_db.hora_fora_estado = hora_fora_estado
        comissao_db.motivo = motivo
        db.session.commit()

        flash('Comissionamento atualizado!', 'success')
        return redirect(url_for('listar_comissionamentos'))

    return render_template('comissionamentos/editar_comissionamento.html',
                          comissao=comissao_db, colaboradores=colaboradores)

@app.route('/deletar-comissionamento/<int:cid>', methods=['POST'])
@login_required
@gestor_required
def deletar_comissionamento(cid):
    """Deleta comissionamento manual"""
    comissao_db = ComissionamentoDB.query.get_or_404(cid)
    cliente = comissao_db.cliente
    db.session.delete(comissao_db)
    db.session.commit()

    flash(f'Comissionamento "{cliente}" deletado!', 'success')
    return redirect(url_for('listar_comissionamentos'))

@app.route('/comissionamentos/exportar-excel')
@login_required
@gestor_required
def exportar_comissionamentos_excel():
    """Exporta comissionamentos agrupados por colaborador para Excel"""
    comissions = ComissionamentoDB.query.all()

    # Agrupar por colaborador
    por_colaborador = {}
    for com in comissions:
        nome_colab = com.consultor.nome if com.consultor else 'Sem Atribuição'
        if nome_colab not in por_colaborador:
            por_colaborador[nome_colab] = {
                'registros': [],
                'total_horas': 0,
                'total_registros': 0
            }
        por_colaborador[nome_colab]['registros'].append(com)
        por_colaborador[nome_colab]['total_horas'] += com.horas_comissionadas
        por_colaborador[nome_colab]['total_registros'] += 1

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comissionamentos'

    # Estilos
    header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    total_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
    total_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalho
    headers = ['Colaborador', 'Cliente', 'Data', 'Horas', 'Hora Fora', 'Motivo']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Dados por colaborador
    row = 2
    geral_total_horas = 0
    geral_total_registros = 0

    for nome_colab in sorted(por_colaborador.keys()):
        grupo = por_colaborador[nome_colab]
        primeiro = True

        for com in grupo['registros']:
            ws.cell(row=row, column=1).value = nome_colab if primeiro else ''
            ws.cell(row=row, column=2).value = com.cliente
            ws.cell(row=row, column=3).value = com.data_comissao.strftime('%d/%m/%Y')
            ws.cell(row=row, column=4).value = com.horas_comissionadas
            ws.cell(row=row, column=5).value = com.hora_fora_estado or ''
            ws.cell(row=row, column=6).value = com.motivo or ''

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')

            primeiro = False
            row += 1

        # Subtotal do colaborador
        ws.cell(row=row, column=1).value = f'SUBTOTAL {nome_colab}'
        ws.cell(row=row, column=4).value = grupo['total_horas']
        ws.cell(row=row, column=2).value = f"{grupo['total_registros']} registro(s)"

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = total_fill
            ws.cell(row=row, column=col).font = total_font
            ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')

        geral_total_horas += grupo['total_horas']
        geral_total_registros += grupo['total_registros']
        row += 2

    # Total geral
    ws.cell(row=row, column=1).value = 'TOTAL GERAL'
    ws.cell(row=row, column=2).value = f'{geral_total_registros} registro(s)'
    ws.cell(row=row, column=4).value = geral_total_horas

    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
        ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFF', size=11)
        ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar arquivo
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'comissionamentos_{date.today().strftime("%Y%m%d")}.xlsx'
    )

# ─── Rotas: projetos ERP ──────────────────────────────────────────────────────

@app.route('/projetos')
@login_required
def listar_projetos():
    """Lista todos os projetos ERP"""
    projetos_db = ERPProjetoDB.query.order_by(ERPProjetoDB.criado_em.desc()).all()
    projetos = [db_to_projeto(p) for p in projetos_db]

    # Filtros
    status_filtro = request.args.get('status', 'todos')
    if status_filtro != 'todos':
        projetos = [p for p in projetos if p.status == status_filtro]

    analytics = ProjetoAnalytics(projetos)
    resumo = analytics.resumo_geral()

    return render_template('projetos/lista_projetos.html',
                          projetos=projetos,
                          status_filtro=status_filtro,
                          resumo=resumo)

@app.route('/dashboard-projetos')
@login_required
def dashboard_projetos():
    """Dashboard com visão geral dos projetos e indicadores por coordenador"""
    projetos_db = ERPProjetoDB.query.all()
    projetos = [db_to_projeto(p) for p in projetos_db]

    analytics = ProjetoAnalytics(projetos)

    # Stats gerais
    stats = {
        'total_projetos': len(projetos),
        'em_andamento': sum(1 for p in projetos if p.status == 'Em andamento'),
        'finalizados': sum(1 for p in projetos if p.status == 'Finalizado'),
        'atrasados': sum(1 for p in projetos if p.esta_atrasado()),
        'progresso_medio': analytics.percentual_conclusao_geral(),
        'valor_total_em_andamento': sum(p.valor_mensalidades for p in projetos if p.status == 'Em andamento')
    }

    # Projetos atrasados
    projetos_atrasados = [p for p in projetos if p.esta_atrasado()]

    # Próximos vencimentos (30 dias)
    projetos_vencimento_proximo = [p for p in projetos
                                   if 0 < p.dias_restantes() <= 30
                                   and p.status == 'Em andamento']
    projetos_vencimento_proximo.sort(key=lambda p: p.dias_restantes())

    # Resumo por coordenador
    coordenadores_resumo = {}
    for p in projetos:
        resp = p.responsavel or 'Sem atribuição'
        if resp not in coordenadores_resumo:
            coordenadores_resumo[resp] = {
                'nome': resp,
                'projetos_total': 0,
                'em_andamento': 0,
                'finalizados': 0,
                'atrasados': 0,
                'valor_total': 0,
                'projetos_lista': []
            }
        coordenadores_resumo[resp]['projetos_total'] += 1
        coordenadores_resumo[resp]['projetos_lista'].append(p)
        if p.status == 'Em andamento':
            coordenadores_resumo[resp]['em_andamento'] += 1
        elif p.status == 'Finalizado':
            coordenadores_resumo[resp]['finalizados'] += 1
        if p.esta_atrasado():
            coordenadores_resumo[resp]['atrasados'] += 1
        coordenadores_resumo[resp]['valor_total'] += p.valor_mensalidades

    coordenadores_resumo = list(coordenadores_resumo.values())
    coordenadores_resumo.sort(key=lambda x: x['projetos_total'], reverse=True)

    # Resumo por potencial
    resumo_potencial = {}
    for p in projetos:
        pot = p.potencial_cliente or 'Não especificado'
        resumo_potencial[pot] = resumo_potencial.get(pot, 0) + 1

    return render_template('projetos/dashboard_projetos.html',
                          stats=stats,
                          projetos_atrasados=projetos_atrasados,
                          projetos_vencimento_proximo=projetos_vencimento_proximo,
                          coordenadores_resumo=coordenadores_resumo,
                          resumo_potencial=resumo_potencial)

@app.route('/novo-projeto', methods=['GET', 'POST'])
@login_required
@gestor_required
def novo_projeto():
    """Cria novo projeto ERP"""
    colaboradores = obter_coordenadores()

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        try:
            data_aceite = datetime.strptime(request.form.get('data_aceite', ''), '%Y-%m-%d').date()
        except ValueError:
            flash('Data de aceite inválida.', 'danger')
            return render_template('projetos/novo_projeto.html', colaboradores=colaboradores)

        # Data de conclusão é opcional
        data_conclusao_str = request.form.get('data_conclusao', '').strip()
        data_conclusao = None
        if data_conclusao_str:
            try:
                data_conclusao = datetime.strptime(data_conclusao_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Data de conclusão inválida.', 'danger')
                return render_template('projetos/novo_projeto.html', colaboradores=colaboradores)

        valor_str = request.form.get('valor_mensalidades', '0').replace(',', '.')
        try:
            valor = float(valor_str) if valor_str else 0
        except ValueError:
            valor = 0

        responsavel_id = request.form.get('responsavel_id')
        if responsavel_id:
            responsavel_id = int(responsavel_id)
        else:
            responsavel_id = None

        descricao = request.form.get('descricao', '').strip()
        numero_unidades = request.form.get('numero_unidades', '1')
        try:
            numero_unidades = int(numero_unidades) if numero_unidades else 1
        except ValueError:
            numero_unidades = 1

        potencial_cliente = request.form.get('potencial_cliente', 'Médio')
        tipo_projeto = request.form.get('tipo_projeto', 'Novo')

        # Validação básica (sem exigir data_conclusao)
        if not nome:
            flash('Nome do projeto é obrigatório.', 'danger')
            return render_template('projetos/novo_projeto.html', colaboradores=colaboradores)

        # Verificar duplicata
        if ERPProjetoDB.query.filter_by(nome_projeto=nome).first():
            flash(f'Projeto com nome "{nome}" já existe.', 'danger')
            return render_template('projetos/novo_projeto.html', colaboradores=colaboradores)

        # Criar novo projeto
        novo = ERPProjetoDB(
            nome_projeto=nome,
            data_aceite=data_aceite,
            data_conclusao=data_conclusao,
            status='Em andamento',
            responsavel_id=responsavel_id,
            valor_mensalidades=valor,
            descricao=descricao,
            percentual_conclusao=0,
            numero_unidades=numero_unidades,
            potencial_cliente=potencial_cliente,
            tipo_projeto=tipo_projeto
        )
        db.session.add(novo)
        db.session.commit()

        flash(f'Projeto "{nome}" criado com sucesso!', 'success')
        return redirect(url_for('detalhe_projeto', pid=novo.id))

    return render_template('projetos/novo_projeto.html', colaboradores=colaboradores)

@app.route('/projeto/<int:pid>')
@login_required
def detalhe_projeto(pid):
    """Exibe detalhes completos de um projeto"""
    p_db = ERPProjetoDB.query.get_or_404(pid)
    projeto = db_to_projeto(p_db)

    colaboradores = ColaboradorDB.query.filter_by(ativo=True).order_by(ColaboradorDB.nome).all()

    return render_template('projetos/detalhe_projeto.html',
                          projeto=projeto,
                          p_db=p_db,
                          colaboradores=colaboradores)

@app.route('/editar-projeto/<int:pid>', methods=['GET', 'POST'])
@login_required
@gestor_required
def editar_projeto(pid):
    """Edita um projeto ERP"""
    p_db = ERPProjetoDB.query.get_or_404(pid)
    colaboradores = obter_coordenadores()

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        try:
            data_aceite = datetime.strptime(request.form.get('data_aceite', ''), '%Y-%m-%d').date()
        except ValueError:
            flash('Data de aceite inválida.', 'danger')
            return redirect(url_for('editar_projeto', pid=pid))

        # Data de conclusão é opcional
        data_conclusao_str = request.form.get('data_conclusao', '').strip()
        data_conclusao = p_db.data_conclusao  # mantém valor atual se vazio
        if data_conclusao_str:
            try:
                data_conclusao = datetime.strptime(data_conclusao_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Data de conclusão inválida.', 'danger')
                return redirect(url_for('editar_projeto', pid=pid))

        valor_str = request.form.get('valor_mensalidades', '0').replace(',', '.')
        try:
            valor = float(valor_str) if valor_str else 0
        except ValueError:
            valor = 0

        status = request.form.get('status', 'Em andamento')
        responsavel_id = request.form.get('responsavel_id')
        if responsavel_id:
            responsavel_id = int(responsavel_id)
        else:
            responsavel_id = None

        descricao = request.form.get('descricao', '').strip()
        percentual = request.form.get('percentual_conclusao', '0').replace(',', '.')
        try:
            percentual = float(percentual) if percentual else 0
        except ValueError:
            percentual = 0

        numero_unidades = request.form.get('numero_unidades', '1')
        try:
            numero_unidades = int(numero_unidades) if numero_unidades else 1
        except ValueError:
            numero_unidades = 1

        potencial_cliente = request.form.get('potencial_cliente', 'Médio')
        tipo_projeto = request.form.get('tipo_projeto', 'Novo')

        # Validação
        valido, erro = ProjetoValidator.validar_projeto(nome, data_aceite, data_conclusao, valor)
        if not valido:
            flash(f'Erro: {erro}', 'danger')
            return redirect(url_for('editar_projeto', pid=pid))

        # Atualizar
        p_db.nome_projeto = nome
        p_db.data_aceite = data_aceite
        p_db.data_conclusao = data_conclusao
        p_db.status = status
        p_db.responsavel_id = responsavel_id
        p_db.valor_mensalidades = valor
        p_db.descricao = descricao
        p_db.percentual_conclusao = percentual
        p_db.numero_unidades = numero_unidades
        p_db.potencial_cliente = potencial_cliente
        p_db.tipo_projeto = tipo_projeto
        db.session.commit()

        flash(f'Projeto "{nome}" atualizado!', 'success')
        return redirect(url_for('detalhe_projeto', pid=pid))

    projeto = db_to_projeto(p_db)
    return render_template('projetos/editar_projeto.html',
                          projeto=projeto,
                          p_db=p_db,
                          colaboradores=colaboradores)

@app.route('/projeto/<int:pid>/concluir', methods=['POST'])
@login_required
@gestor_required
def concluir_projeto(pid):
    """Marca projeto como finalizado"""
    p_db = ERPProjetoDB.query.get_or_404(pid)

    if p_db.status == 'Finalizado':
        flash('Projeto já está finalizado.', 'warning')
    else:
        p_db.status = 'Finalizado'
        p_db.percentual_conclusao = 100
        db.session.commit()
        flash(f'Projeto "{p_db.nome_projeto}" marcado como finalizado!', 'success')

    return redirect(url_for('detalhe_projeto', pid=pid))

@app.route('/projeto/<int:pid>/reativar', methods=['POST'])
@login_required
@gestor_required
def reativar_projeto(pid):
    """Volta projeto de Finalizado para Em andamento"""
    p_db = ERPProjetoDB.query.get_or_404(pid)

    if p_db.status != 'Finalizado':
        flash('Apenas projetos finalizados podem ser reativados.', 'warning')
    else:
        p_db.status = 'Em andamento'
        db.session.commit()
        flash(f'Projeto "{p_db.nome_projeto}" reativado para Em Andamento!', 'success')

    return redirect(url_for('detalhe_projeto', pid=pid))

@app.route('/projeto/<int:pid>/modulo', methods=['POST'])
@login_required
@gestor_required
def adicionar_modulo(pid):
    """Adiciona módulo a um projeto"""
    p_db = ERPProjetoDB.query.get_or_404(pid)

    nome = request.form.get('modulo', '').strip()
    percentual_str = request.form.get('percentual_conclusao', '0').replace(',', '.')

    # Validação do nome
    if not nome:
        flash('Nome do módulo é obrigatório.', 'danger')
        return redirect(url_for('detalhe_projeto', pid=pid))

    # Processar percentual
    try:
        percentual = float(percentual_str)
        percentual = max(0, min(100, percentual))  # Limitar entre 0 e 100
    except ValueError:
        percentual = 0

    novo_modulo = ERPModuloDB(
        projeto_id=pid,
        modulo=nome,
        status_modulo='Planejado',
        data_inicio_modulo=None,
        data_conclusao_modulo=None,
        percentual_conclusao=percentual
    )
    db.session.add(novo_modulo)
    db.session.commit()

    flash(f'Módulo "{nome}" adicionado!', 'success')
    return redirect(url_for('detalhe_projeto', pid=pid))

@app.route('/projeto/<int:pid>/unidade', methods=['POST'])
@login_required
@gestor_required
def adicionar_unidade(pid):
    """Adiciona unidade a um projeto"""
    p_db = ERPProjetoDB.query.get_or_404(pid)

    nome = request.form.get('unidade', '').strip()

    # Validação
    valido, erro = ProjetoValidator.validar_unidade(nome)
    if not valido:
        flash(f'Erro: {erro}', 'danger')
        return redirect(url_for('detalhe_projeto', pid=pid))

    nova_unidade = ERPUnidadeDB(
        projeto_id=pid,
        unidade=nome,
        status_unidade='Não iniciado',
        data_inicio_unidade=None,
        data_conclusao_unidade=None
    )
    db.session.add(nova_unidade)
    db.session.commit()

    flash(f'Unidade "{nome}" adicionada!', 'success')
    return redirect(url_for('detalhe_projeto', pid=pid))

@app.route('/projeto/<int:mid>/editar-modulo/<int:mid_id>', methods=['POST'])
@login_required
@gestor_required
def editar_modulo(mid, mid_id):
    """Edita um módulo"""
    modulo_db = ERPModuloDB.query.get_or_404(mid_id)

    # Permitir alterar nome do módulo
    nome = request.form.get('modulo', modulo_db.modulo).strip()
    status = request.form.get('status', modulo_db.status_modulo)
    percentual_str = request.form.get('percentual_conclusao', str(modulo_db.percentual_conclusao or 0)).replace(',', '.')

    try:
        percentual = float(percentual_str)
        percentual = max(0, min(100, percentual))  # Limitar entre 0 e 100
    except ValueError:
        percentual = modulo_db.percentual_conclusao or 0

    modulo_db.modulo = nome
    modulo_db.status_modulo = status
    modulo_db.percentual_conclusao = percentual
    db.session.commit()

    flash('Módulo atualizado!', 'success')
    return redirect(url_for('detalhe_projeto', pid=mid))

@app.route('/projeto/<int:pid>/editar-unidade/<int:uid>', methods=['POST'])
@login_required
@gestor_required
def editar_unidade(pid, uid):
    """Edita uma unidade"""
    unidade_db = ERPUnidadeDB.query.get_or_404(uid)

    status = request.form.get('status', unidade_db.status_unidade)

    unidade_db.status_unidade = status
    db.session.commit()

    flash('Unidade atualizada!', 'success')
    return redirect(url_for('detalhe_projeto', pid=pid))

@app.route('/projeto/<int:pid>/atividade', methods=['POST'])
@login_required
@gestor_required
def adicionar_atividade(pid):
    """Adiciona atividade a um projeto"""
    p_db = ERPProjetoDB.query.get_or_404(pid)

    titulo = request.form.get('titulo', '').strip()
    if not titulo:
        flash('Título da atividade é obrigatório.', 'danger')
        return redirect(url_for('detalhe_projeto', pid=pid))

    nova_atividade = ERPAtividadeDB(
        projeto_id=pid,
        titulo=titulo,
        data_reuniao=date.today(),
        status_atividade='Aberta',
        concluida=False
    )
    db.session.add(nova_atividade)
    db.session.commit()

    flash(f'Atividade "{titulo}" adicionada!', 'success')
    return redirect(url_for('detalhe_projeto', pid=pid))

@app.route('/projeto/<int:pid>/deletar-atividade/<int:aid>', methods=['POST'])
@login_required
@gestor_required
def deletar_atividade(pid, aid):
    """Deleta uma atividade"""
    atividade_db = ERPAtividadeDB.query.get_or_404(aid)
    titulo = atividade_db.titulo

    db.session.delete(atividade_db)
    db.session.commit()

    flash(f'Atividade "{titulo}" deletada!', 'success')
    return redirect(url_for('detalhe_projeto', pid=pid))

# ─── API: saldo ───────────────────────────────────────────────────────────────

@app.route('/api/saldo/<int:cid>')
@login_required
def api_saldo(cid):
    colaboradores_raw, ferias_plan, ferias_real = carregar_tudo()
    c = next((x for x in colaboradores_raw if x.id == cid), None)
    if not c:
        return jsonify({'erro': 'não encontrado'}), 404
    s_raw = saldo_colab(c, ferias_real, ferias_plan)
    if s_raw is None:
        return jsonify({'saldo': None, 'status': 'estagiario', 'colaborador': c.nome, 'estagiario': True})
    s_disp = saldo_quantizado(s_raw)
    return jsonify({'saldo': s_disp, 'saldo_raw': s_raw, 'status': status_saldo(s_raw), 'colaborador': c.nome})

@app.route('/api/projetos')
@login_required
def api_projetos():
    """API JSON com status geral dos projetos"""
    projetos_db = ERPProjetoDB.query.all()
    projetos = [db_to_projeto(p) for p in projetos_db]

    analytics = ProjetoAnalytics(projetos)
    resumo = analytics.resumo_geral()
    criticos = analytics.projetos_criticos()

    return jsonify({
        'resumo': resumo,
        'projetos_criticos': [
            {
                'id': c['projeto'].id,
                'nome': c['projeto'].nome,
                'status': c['projeto'].status,
                'motivos': c['motivos'],
                'severidade': c['severidade']
            }
            for c in criticos
        ]
    })

# ─── Inicialização do Banco (Vercel) ───────────────────────────────────────────

@app.route('/init', methods=['GET'])
def init_database():
    """
    Endpoint para inicializar o banco de dados no Vercel.
    Acesse https://seu-app.vercel.app/init uma única vez para criar as tabelas.
    """
    try:
        with app.app_context():
            # Criar todas as tabelas
            db.create_all()

            # Verificar se usuário admin existe
            admin = User.query.filter_by(username='admin').first()

            if not admin:
                # Criar usuário admin
                admin = User(
                    username=os.environ.get('ADMIN_USER', 'admin'),
                    nome='Administrador',
                    perfil='gestor',
                    ativo=True
                )
                admin.set_senha(os.environ.get('ADMIN_PASSWORD', 'admin123'))
                db.session.add(admin)
                db.session.commit()

                return jsonify({
                    'status': 'success',
                    'message': '✅ Banco de dados inicializado com sucesso!',
                    'credentials': {
                        'username': 'admin',
                        'password': 'admin123'
                    }
                }), 200
            else:
                return jsonify({
                    'status': 'info',
                    'message': '✅ Banco de dados já estava inicializado!'
                }), 200

    except Exception as e:
        # Se falhar com Supabase, tentar com SQLite como fallback
        try:
            print(f"⚠️ Supabase falhou, usando SQLite como fallback: {str(e)}", file=sys.stderr)

            # Mudar para SQLite
            app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(DATA_DIR, "ferias_data.db")}'

            with app.app_context():
                # Criar todas as tabelas no SQLite
                db.create_all()

                # Criar usuário admin se não existir
                admin = User.query.filter_by(username='admin').first()
                if not admin:
                    admin = User(
                        username='admin',
                        nome='Administrador',
                        perfil='gestor',
                        ativo=True
                    )
                    admin.set_senha('admin123')
                    db.session.add(admin)
                    db.session.commit()

                return jsonify({
                    'status': 'success_sqlite',
                    'message': '✅ Banco inicializado com SQLite (fallback)!',
                    'warning': 'Usando SQLite localmente. Para usar Supabase, configure DATABASE_URL.',
                    'credentials': {
                        'username': 'admin',
                        'password': 'admin123'
                    }
                }), 200
        except Exception as fallback_error:
            return jsonify({
                'status': 'error',
                'message': f'❌ Erro: {str(fallback_error)}',
                'original_error': str(e)
            }), 500

# ─── Migração de Dados (Local → Vercel) ────────────────────────────────────────

@app.route('/migrate-data', methods=['POST'])
def migrate_data():
    """
    Endpoint para receber dados do banco local e inserir no banco online.
    Usado pelo script migrate_local_to_vercel.py
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({'status': 'error', 'message': 'Nenhum dado fornecido'}), 400

        summary = {}

        with app.app_context():
            # Migrar colaboradores
            if 'colaboradores' in data:
                for row in data['colaboradores']:
                    colab = ColaboradorDB(
                        id=row.get('id'),
                        nome=row.get('nome'),
                        data_admissao=row.get('data_admissao'),
                        time=row.get('time', ''),
                        cidade=row.get('cidade', ''),
                        ativo=bool(row.get('ativo', True))
                    )
                    db.session.merge(colab)
                db.session.commit()
                summary['colaboradores'] = len(data['colaboradores'])

            # Migrar usuários
            if 'users' in data:
                for row in data['users']:
                    user = User(
                        id=row.get('id'),
                        username=row.get('username'),
                        nome=row.get('nome'),
                        senha_hash=row.get('senha_hash'),
                        perfil=row.get('perfil', 'colaborador'),
                        colaborador_id=row.get('colaborador_id'),
                        ativo=bool(row.get('ativo', True))
                    )
                    db.session.merge(user)
                db.session.commit()
                summary['users'] = len(data['users'])

            # Migrar férias
            if 'ferias' in data:
                for row in data['ferias']:
                    feria = FeriasDB(
                        id=row.get('id'),
                        colaborador_id=row.get('colaborador_id'),
                        data_inicio=row.get('data_inicio'),
                        data_fim=row.get('data_fim'),
                        dias=row.get('dias'),
                        status=row.get('status', 'Planejado'),
                        conflito_detectado=bool(row.get('conflito_detectado', False)),
                        conflito_aprovado=bool(row.get('conflito_aprovado', False))
                    )
                    db.session.merge(feria)
                db.session.commit()
                summary['ferias'] = len(data['ferias'])

            # Migrar projetos ERP
            if 'erp_projetos' in data:
                for row in data['erp_projetos']:
                    proj = ERPProjetoDB(
                        id=row.get('id'),
                        nome_projeto=row.get('nome_projeto'),
                        data_aceite=row.get('data_aceite'),
                        data_conclusao=row.get('data_conclusao'),
                        status=row.get('status', 'Em andamento'),
                        responsavel_id=row.get('responsavel_id'),
                        valor_mensalidades=row.get('valor_mensalidades', 0),
                        descricao=row.get('descricao', ''),
                        percentual_conclusao=row.get('percentual_conclusao', 0),
                        numero_unidades=row.get('numero_unidades', 1),
                        potencial_cliente=row.get('potencial_cliente', 'Médio'),
                        tipo_projeto=row.get('tipo_projeto', 'Novo'),
                        criado_em=row.get('criado_em'),
                        atualizado_em=row.get('atualizado_em')
                    )
                    db.session.merge(proj)
                db.session.commit()
                summary['erp_projetos'] = len(data['erp_projetos'])

            # Migrar módulos
            if 'erp_modulos' in data:
                for row in data['erp_modulos']:
                    mod = ERPModuloDB(
                        id=row.get('id'),
                        projeto_id=row.get('projeto_id'),
                        modulo=row.get('modulo'),
                        status_modulo=row.get('status_modulo', 'Planejado'),
                        data_inicio_modulo=row.get('data_inicio_modulo'),
                        data_conclusao_modulo=row.get('data_conclusao_modulo'),
                        percentual_conclusao=row.get('percentual_conclusao', 0),
                        criado_em=row.get('criado_em')
                    )
                    db.session.merge(mod)
                db.session.commit()
                summary['erp_modulos'] = len(data['erp_modulos'])

            # Migrar unidades
            if 'erp_unidades' in data:
                for row in data['erp_unidades']:
                    uni = ERPUnidadeDB(
                        id=row.get('id'),
                        projeto_id=row.get('projeto_id'),
                        unidade=row.get('unidade'),
                        status_unidade=row.get('status_unidade', 'Não iniciado'),
                        data_inicio_unidade=row.get('data_inicio_unidade'),
                        data_conclusao_unidade=row.get('data_conclusao_unidade'),
                        criado_em=row.get('criado_em')
                    )
                    db.session.merge(uni)
                db.session.commit()
                summary['erp_unidades'] = len(data['erp_unidades'])

            # Migrar atividades
            if 'erp_atividades' in data:
                for row in data['erp_atividades']:
                    ativ = ERPAtividadeDB(
                        id=row.get('id'),
                        projeto_id=row.get('projeto_id'),
                        titulo=row.get('titulo'),
                        descricao=row.get('descricao', ''),
                        data_reuniao=row.get('data_reuniao'),
                        responsavel_nota=row.get('responsavel_nota', ''),
                        status_atividade=row.get('status_atividade', 'Aberta'),
                        concluida=bool(row.get('concluida', False)),
                        criado_em=row.get('criado_em'),
                        atualizado_em=row.get('atualizado_em')
                    )
                    db.session.merge(ativ)
                db.session.commit()
                summary['erp_atividades'] = len(data['erp_atividades'])

            # Migrar comissionamentos
            if 'comissionamentos' in data:
                for row in data['comissionamentos']:
                    comis = ComissionamentoDB(
                        id=row.get('id'),
                        consultor_id=row.get('consultor_id'),
                        cliente=row.get('cliente'),
                        data_comissao=row.get('data_comissao'),
                        horas_comissionadas=row.get('horas_comissionadas'),
                        hora_fora_estado=row.get('hora_fora_estado', ''),
                        motivo=row.get('motivo', ''),
                        periodo_inicio=row.get('periodo_inicio'),
                        periodo_fim=row.get('periodo_fim'),
                        criado_em=row.get('criado_em'),
                        atualizado_em=row.get('atualizado_em')
                    )
                    db.session.merge(comis)
                db.session.commit()
                summary['comissionamentos'] = len(data['comissionamentos'])

        return jsonify({
            'status': 'success',
            'message': '✅ Dados migrados com sucesso para o Vercel!',
            'summary': summary
        }), 200

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'❌ Erro na migração: {str(e)}'
        }), 500

# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
