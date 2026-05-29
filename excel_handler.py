import os
from datetime import datetime
from typing import List, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import Colaborador, Ferias


class ExcelHandler:
    """Gerencia leitura e escrita de dados em Excel"""

    def __init__(self, caminho_arquivo: str = "data/ferias_data.xlsx"):
        self.caminho = caminho_arquivo
        self.workbook = None
        self._garantir_arquivo_existe()

    def _garantir_arquivo_existe(self):
        """Cria arquivo Excel com abas se não existir"""
        os.makedirs(os.path.dirname(self.caminho), exist_ok=True)

        if not os.path.exists(self.caminho):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove aba padrão

            # Aba: Colaboradores
            ws_colab = wb.create_sheet("Colaboradores")
            ws_colab.append(["ID", "Nome", "Data Admissão", "Departamento/Time", "Ativo", "Cidade"])
            ws_colab.column_dimensions['A'].width = 5
            ws_colab.column_dimensions['B'].width = 25
            ws_colab.column_dimensions['C'].width = 18
            ws_colab.column_dimensions['D'].width = 20
            ws_colab.column_dimensions['E'].width = 10
            ws_colab.column_dimensions['F'].width = 20

            # Aba: Férias Planejadas
            ws_plan = wb.create_sheet("Férias Planejadas")
            ws_plan.append(["ID", "ID_Collab", "Nome", "Início", "Fim", "Dias", "Status",
                           "Conflito Detectado", "Conflito Aprovado", "Observações"])
            for col in ws_plan.columns:
                ws_plan.column_dimensions[col[0].column_letter].width = 18

            # Aba: Férias Realizadas
            ws_real = wb.create_sheet("Férias Realizadas")
            ws_real.append(["ID", "ID_Collab", "Nome", "Início", "Fim", "Dias", "Ano"])
            for col in ws_real.columns:
                ws_real.column_dimensions[col[0].column_letter].width = 15

            wb.save(self.caminho)

    def _migrar_coluna_cidade(self):
        """Adiciona coluna Cidade se não existir (migração automática)."""
        wb = openpyxl.load_workbook(self.caminho)
        ws = wb["Colaboradores"]
        cabecalho = [cell.value for cell in ws[1]]
        if "Cidade" not in cabecalho:
            col_idx = len(cabecalho) + 1
            ws.cell(row=1, column=col_idx, value="Cidade")
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
            wb.save(self.caminho)
        wb.close()

    def carregar_colaboradores(self) -> List[Colaborador]:
        """Carrega todos os colaboradores do Excel"""
        self._migrar_coluna_cidade()
        wb = openpyxl.load_workbook(self.caminho)
        ws = wb["Colaboradores"]

        # Mapeia colunas pelo cabeçalho (tolerante a ordem)
        cabecalho = {cell.value: cell.column - 1 for cell in ws[1] if cell.value}

        colaboradores = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            id_val  = int(row[cabecalho.get('ID', 0)])
            nome    = row[cabecalho.get('Nome', 1)]
            data_ad = row[cabecalho.get('Data Admissão', 2)]
            time_   = row[cabecalho.get('Departamento/Time', 3)]
            ativo   = str(row[cabecalho.get('Ativo', 4)] or 'Sim').lower() == 'sim'
            cidade  = str(row[cabecalho.get('Cidade', 5)] or '') if 'Cidade' in cabecalho else ''

            if isinstance(data_ad, str):
                data_ad = datetime.strptime(data_ad, "%Y-%m-%d")

            colaboradores.append(Colaborador(id_val, nome, data_ad, time_, ativo, cidade))

        wb.close()
        return colaboradores

    def salvar_colaboradores(self, colaboradores: List[Colaborador]):
        """Salva colaboradores no Excel"""
        wb = openpyxl.load_workbook(self.caminho)
        ws = wb["Colaboradores"]

        # Limpa dados existentes (mantém header)
        ws.delete_rows(2, ws.max_row)

        # Escreve novos dados
        for colab in colaboradores:
            ws.append([
                colab.id,
                colab.nome,
                colab.data_admissao.strftime("%Y-%m-%d"),
                colab.time,
                "Sim" if colab.ativo else "Não",
                getattr(colab, 'cidade', '') or '',
            ])

        wb.save(self.caminho)
        wb.close()

    def carregar_ferias_planejadas(self) -> List[Ferias]:
        """Carrega férias planejadas do Excel"""
        wb = openpyxl.load_workbook(self.caminho)
        ws = wb["Férias Planejadas"]

        ferias = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value is None:
                break

            id_val = int(row[0].value)
            colaborador_id = int(row[1].value)
            data_inicio = row[3].value
            data_fim = row[4].value
            status = row[6].value or "Planejado"
            conflito_detectado = str(row[7].value or "Não").lower() == 'sim'
            conflito_aprovado = str(row[8].value or "Não").lower() == 'sim'

            if isinstance(data_inicio, str):
                data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
            if isinstance(data_fim, str):
                data_fim = datetime.strptime(data_fim, "%Y-%m-%d")

            f = Ferias(id_val, colaborador_id, data_inicio, data_fim, status, conflito_detectado, conflito_aprovado)
            ferias.append(f)

        wb.close()
        return ferias

    def salvar_ferias_planejadas(self, ferias_list: List[Ferias]):
        """Salva férias planejadas no Excel"""
        wb = openpyxl.load_workbook(self.caminho)
        ws = wb["Férias Planejadas"]

        # Limpa dados existentes
        ws.delete_rows(2, ws.max_row)

        # Escreve novos dados
        for f in ferias_list:
            ws.append([
                f.id,
                f.colaborador_id,
                "",  # Nome será preenchido pela app
                f.data_inicio.strftime("%Y-%m-%d"),
                f.data_fim.strftime("%Y-%m-%d"),
                f.dias,
                f.status,
                "Sim" if f.conflito_detectado else "Não",
                "Sim" if f.conflito_aprovado else "Não",
                ""  # Observações
            ])

        wb.save(self.caminho)
        wb.close()

    def salvar_ferias_realizadas(self, ferias: Ferias, colaborador_nome: str):
        """Copia uma férias para o histórico de realizadas"""
        wb = openpyxl.load_workbook(self.caminho)
        ws = wb["Férias Realizadas"]

        ano = ferias.data_inicio.year
        proximo_id = (ws.max_row or 1)

        ws.append([
            proximo_id,
            ferias.colaborador_id,
            colaborador_nome,
            ferias.data_inicio.strftime("%Y-%m-%d"),
            ferias.data_fim.strftime("%Y-%m-%d"),
            ferias.dias,
            ano
        ])

        wb.save(self.caminho)
        wb.close()

    def obter_proximo_id_ferias(self) -> int:
        """Obtém o próximo ID disponível para férias"""
        wb = openpyxl.load_workbook(self.caminho)
        ws = wb["Férias Planejadas"]

        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                max_id = max(max_id, int(row[0]))

        wb.close()
        return max_id + 1

    def obter_proximo_id_colaborador(self) -> int:
        """Obtém o próximo ID disponível para colaborador"""
        wb = openpyxl.load_workbook(self.caminho)
        ws = wb["Colaboradores"]

        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                max_id = max(max_id, int(row[0]))

        wb.close()
        return max_id + 1
